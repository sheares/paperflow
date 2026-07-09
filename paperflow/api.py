"""paperflow API: serves the trust UI, pile views shaped for it, and the
chat router. Route (b): the verified mockup is the front end.

If FIREWORKS_API_KEY is absent, everything runs full-local (zero egress);
the UI's receipts reflect whichever route actually executed.
"""
from __future__ import annotations

import asyncio
import json
import os
import shutil
import tempfile
import uuid
from pathlib import Path

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from pydantic import BaseModel

from .extractor import extract_pile
from .models import profile as _model_profile, resolve as _resolve_model
from .pipeline import run_pile
from .router import Router
from .uiview import build_pile_view

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "outputs"
PILES = {
    "kyc_onboarding": "kyc",
    "patient_intake": "patient",
    "partner_collation": "partner",
}

app = FastAPI(title="paperflow")
_routers: dict[str, Router] = {}

# session-scoped "real" piles: session_id -> {"dir": Path, "schema": Path}
REAL = Path(tempfile.gettempdir()) / "paperflow_real"
REAL.mkdir(exist_ok=True)
_real_sessions: dict[str, dict] = {}
ALLOWED_UPLOAD_SUFFIXES = {".pdf", ".txt", ".md", ".xlsx"}
MAX_UPLOAD_BYTES = 25 * 1024 * 1024   # real KYC/legal PDFs can be sizeable
MAX_UPLOAD_FILES = 12


def _schema(pile: str) -> Path:
    return ROOT / "paperflow" / "schemas" / f"{PILES[pile]}.yaml"


def _ensure_run(pile: str, full_local: bool | None = None) -> Path:
    if pile not in PILES:
        raise HTTPException(404, f"unknown pile {pile}")
    run_dir = OUT / f"run_{pile}"
    if not (run_dir / "run_output.json").exists():
        if full_local is None:
            full_local = "FIREWORKS_API_KEY" not in os.environ
        cached = OUT / f"extraction_{pile}.json"
        run_pile(ROOT / "synthetic" / pile, _schema(pile),
                 cached if cached.exists() else None, full_local, OUT)
        _routers.pop(pile, None)
    return run_dir


class AskBody(BaseModel):
    pile: str
    question: str
    full_local: bool = False


class RunBody(BaseModel):
    pile: str
    full_local: bool | None = None


@app.get("/health")
def health():
    return {"status": "ok", "remote_available": "FIREWORKS_API_KEY" in os.environ}


@app.get("/api/status")
def status():
    """Live reachability probe for the header badge. Truthful failure modes:
    local_reachable=false means the MI300X endpoint isn't answering (so
    PDF extraction will fail); remote_configured=false means no
    FIREWORKS_API_KEY (so hybrid asks degrade to full-local)."""
    import httpx
    vllm_url = os.environ.get("VLLM_URL", "")
    local_reachable = False
    vllm_model = None
    if vllm_url:
        try:
            r = httpx.get(f"{vllm_url.rstrip('/')}/models", timeout=2)
            if r.status_code == 200:
                data = r.json().get("data", [])
                local_reachable = bool(data)
                vllm_model = data[0]["id"] if data else None
        except Exception:  # noqa: BLE001
            pass
    return {
        "local_reachable": local_reachable,
        "local_model": vllm_model,
        "local_url": vllm_url,
        "remote_configured": "FIREWORKS_API_KEY" in os.environ,
        "remote_model": _resolve_model(
            os.environ.get("FIREWORKS_MODEL", "deepseek")),
        "remote_model_label": (
            (_model_profile(os.environ.get("FIREWORKS_MODEL", "deepseek")) or {})
            .get("label", "Fireworks AI")),
    }


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    # tiny inline PNG (no external asset needed)
    from fastapi.responses import Response
    return Response(status_code=204)


@app.get("/api/config")
def config():
    """Server-emitted config the client boots off. Single source of
    truth for the routing regex (so JS predictRouting doesn't drift
    from Python LOCAL_TRIGGERS) and the token-family → chip-colour
    map (which was duplicated in three places). Cached client-side
    once at page load."""
    from .router import LOCAL_TRIGGERS
    from .uiview import FAMILY_TYPES
    return {
        # Python re -> JS RegExp: pattern strings are portable for the
        # constructs actually used here (word boundaries, character
        # classes, alternation). Flag 'i' passed separately on the JS
        # side; re.I is already baked into the compiled Python regex.
        "local_triggers": {
            "pattern": LOCAL_TRIGGERS.pattern,
            "flags": "i",
        },
        "family_types": dict(FAMILY_TYPES),
    }


@app.get("/api/classify")
def classify_route(q: str = ""):
    """Ask the SERVER which route a question would take. Client uses
    this on the send-gate + pending-chip path so the prediction is
    the same code path Router.classify() runs at send time — no more
    JS/Python drift possible."""
    from .router import LOCAL_TRIGGERS
    route = "local" if LOCAL_TRIGGERS.search(q or "") else "hybrid"
    return {"route": route}


# Content-Security-Policy backstop for the XSS escape pass. Even if a
# renderer misses an interpolation, an <img onerror=...> or <script>
# from an uploaded doc can't fetch/execute external scripts. Inline
# scripts + styles are still allowed because the app is a single-file
# HTML with heavy inline CSS/JS; connect-src stays 'self' so any
# cloud call has to route through /api/*. Fonts are self-hosted at
# /fonts/ so the entire zero-egress claim holds from page load —
# no fonts.googleapis.com or fonts.gstatic.com in the allowlist.
_CSP = (
    "default-src 'self'; "
    "script-src 'self' 'unsafe-inline'; "
    "style-src 'self' 'unsafe-inline'; "
    "font-src 'self'; "
    "img-src 'self' data:; "
    "frame-src 'self' blob:; "
    "connect-src 'self'; "
    "form-action 'self'; "
    "frame-ancestors 'none'; "
    "base-uri 'none'"
)


@app.get("/")
def index():
    # no-store so UI changes take effect on refresh without a hard reload
    return FileResponse(ROOT / "ui" / "index.html",
                        headers={"Cache-Control": "no-store, must-revalidate",
                                 "Content-Security-Policy": _CSP,
                                 "X-Content-Type-Options": "nosniff",
                                 "Referrer-Policy": "no-referrer"})


@app.get("/fonts/{filename}")
def font_asset(filename: str):
    """Serve self-hosted font files at /fonts/*. Contained to
    ui/fonts/ with a strict allowlist so no path traversal or wild
    filename hits — only the two .woff2 files and the fonts.css
    manifest referenced by ui/index.html."""
    allowed = {"Inter.woff2", "JetBrainsMono.woff2", "fonts.css"}
    if filename not in allowed:
        raise HTTPException(404, "not found")
    path = ROOT / "ui" / "fonts" / filename
    if not path.exists():
        raise HTTPException(404, "not found")
    mime = ("font/woff2" if filename.endswith(".woff2")
            else "text/css")
    return FileResponse(path, media_type=mime,
                        headers={"Cache-Control": "public, max-age=31536000, immutable"})


@app.get("/api/piles")
def piles():
    return {"piles": list(PILES)}


@app.get("/api/doc/{pile}/{filename}")
def get_doc(pile: str, filename: str):
    if pile not in PILES or "/" in filename or ".." in filename:
        raise HTTPException(404, "not found")
    path = ROOT / "synthetic" / pile / filename
    if not path.exists():
        raise HTTPException(404, "not found")
    mime = {"pdf": "application/pdf", "txt": "text/plain",
            "md": "text/markdown",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }.get(path.suffix.lstrip("."), "application/octet-stream")
    from fastapi.responses import FileResponse
    return FileResponse(path, media_type=mime)


def _read_xlsx_preview(path: Path, n: int) -> dict:
    """Shared xlsx-preview payload builder. Used by both the canned
    demo-pile endpoint and the Real-pile endpoint below so their
    shapes stay identical."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
            if len(rows) >= n:
                break
        widths = []
        for i in range(1, (ws.max_column or 1) + 1):
            col = ws.column_dimensions.get(get_column_letter(i))
            widths.append(col.width if col and col.width else 12)
        merges = []
        for m in ws.merged_cells.ranges:
            if m.min_col == 1 and m.max_col >= (ws.max_column or 1) \
                    and m.min_row == m.max_row:
                merges.append(m.min_row - 1)   # 0-indexed row
        sheets.append({"name": ws.title, "rows": rows,
                       "widths": widths, "spanned_rows": merges,
                       "total_rows": ws.max_row, "total_cols": ws.max_column})
    return {"sheets": sheets, "showing": n}


@app.get("/api/xlsx_preview/{pile}/{filename}")
def xlsx_preview(pile: str, filename: str, n: int = 24):
    if pile not in PILES or "/" in filename or ".." in filename:
        raise HTTPException(404, "not found")
    path = ROOT / "synthetic" / pile / filename
    if not path.exists() or path.suffix != ".xlsx":
        raise HTTPException(404, "not found")
    return _read_xlsx_preview(path, n)


class RedactBody(BaseModel):
    text: str


@app.post("/api/redact")
def redact(body: RedactBody):
    """Paste-and-redact sandbox. No pile, no reconciliation: just the
    privacy round-trip on the submitted text, so anyone can watch their
    own words get tokenised. Text stays in memory; no logging, no
    persistence."""
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(400, "empty text")
    if len(text) > 20_000:
        raise HTTPException(413, "text too large (max 20k characters)")
    from paperflow.privacy.redactor import PrivacyRoundTrip
    result = PrivacyRoundTrip().process_pile({"pasted.txt": text})
    emap = result.entity_map
    entities = []
    for tok, canonical in emap.token_to_value.items():
        forms = sorted({emap._display[k] for k, t in emap._lookup.items()
                        if t == tok})
        family = tok.strip("[]").rsplit("_", 1)[0]
        entities.append({"token": tok, "family": family, "forms": forms,
                         "canonical": canonical})
    return {
        "original": text,
        "redacted": result.redacted["pasted.txt"],
        "rehydrated": emap.rehydrate(result.redacted["pasted.txt"]),
        "entities": entities,
        "counts": {"tokens": len(emap.token_to_value),
                   "surface_forms": len(emap._lookup)},
    }


@app.get("/api/pile/{pile}")
def pile_view(pile: str):
    run_dir = _ensure_run(pile)
    extraction = OUT / f"extraction_{pile}.json"
    if not extraction.exists():
        raise HTTPException(409, "no extraction artefact; run a GPU session "
                                 "or POST /api/run first")
    return build_pile_view(run_dir, extraction, _schema(pile))


@app.post("/api/run")
def run(body: RunBody):
    run_dir = _ensure_run(body.pile, body.full_local)
    return {"ok": True, "run_dir": str(run_dir)}


@app.post("/api/ask")
def ask(body: AskBody):
    if body.pile.startswith("real:"):
        sess = _real_session(body.pile[len("real:"):])
        run_dir = sess["dir"] / "run"
        if not (run_dir / "run_output.json").exists():
            raise HTTPException(409, "run the real pile first")
        rk = "real:" + body.pile[len("real:"):]
        if rk not in _routers:
            _routers[rk] = Router(run_dir)
        full_local = body.full_local or "FIREWORKS_API_KEY" not in os.environ
        return _routers[rk].ask(body.question, full_local=full_local)
    run_dir = _ensure_run(body.pile)
    if body.pile not in _routers:
        _routers[body.pile] = Router(run_dir)
    full_local = body.full_local or "FIREWORKS_API_KEY" not in os.environ
    return _routers[body.pile].ask(body.question, full_local=full_local)


# ---------- real pile ----------

def _real_session(session_id: str) -> dict:
    if session_id not in _real_sessions or not \
            _real_sessions[session_id]["dir"].exists():
        raise HTTPException(404, "no such real-pile session")
    return _real_sessions[session_id]


ALLOWED_SCHEMAS = {"kyc", "patient", "partner", "generic"}

# Curated sample piles that the empty-state UI can drop into a fresh
# session with one click. Each maps to a synthetic/<name>/ folder and
# uses the pile-appropriate schema (so the sample loads with domain
# rationale on, not Auto).
SAMPLE_PILES = {
    "individual": {"folder": "kyc_onboarding",  "schema": "kyc",
                   "label": "Individual identity"},
    "corporate":  {"folder": "partner_collation", "schema": "partner",
                   "label": "Corporate identity"},
    "case":       {"folder": "patient_intake",    "schema": "patient",
                   "label": "Case records"},
}


@app.post("/api/real/load_sample/{name}")
def real_load_sample(name: str):
    """One-click load a curated sample pile into a fresh real-pile
    session. Copies the sample docs into a new session dir, sets the
    matching schema, but does NOT run the pipeline — client posts
    /api/real/run separately so the same progress overlay animates."""
    if name not in SAMPLE_PILES:
        raise HTTPException(404, f"no sample named '{name}'")
    src_dir = ROOT / "synthetic" / SAMPLE_PILES[name]["folder"]
    if not src_dir.exists():
        raise HTTPException(500, f"sample folder missing: {src_dir}")
    session_id = uuid.uuid4().hex[:12]
    sess_dir = REAL / session_id
    sess_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for f in sorted(src_dir.iterdir()):
        # skip eval artefacts — samples are demo docs, not scorer input
        if not f.is_file() or f.name == "ground_truth.json":
            continue
        (sess_dir / f.name).write_bytes(f.read_bytes())
        copied.append(f.name)
    # If a cached extraction exists at outputs/extraction_<folder>.json,
    # copy it into the session too. real_run will see the file and skip
    # the live Gemma call — samples then work even with the MI300X
    # asleep. Filenames in the cached JSON match what we just copied
    # (both come from synthetic/<folder>/), so the pipeline treats it
    # as a normal cached run.
    cached_extraction = ROOT / "outputs" / f"extraction_{SAMPLE_PILES[name]['folder']}.json"
    if cached_extraction.exists():
        (sess_dir / "extraction.json").write_bytes(cached_extraction.read_bytes())
    _real_sessions[session_id] = {"dir": sess_dir,
                                  "schema": SAMPLE_PILES[name]["schema"]}
    return {"session_id": session_id,
            "sample": name,
            "label": SAMPLE_PILES[name]["label"],
            "schema": SAMPLE_PILES[name]["schema"],
            "cached": cached_extraction.exists(),
            "files": copied}


@app.post("/api/real/upload")
async def real_upload(schema: str = Form("generic"),
                      session_id: str = Form(""),
                      files: list[UploadFile] = File(...)):
    """Create a fresh session (or add to an existing one) with the uploaded
    documents. Pass session_id to append; leave empty to start fresh.

    Validates ALL files in memory BEFORE writing any to disk — a bad
    file in position 3 of a fresh upload would previously have left
    positions 1-2 as orphans on disk with no session to track them.
    Fixed: buffer + validate first, write only if the whole batch is
    good, clean up cleanly on any error."""
    if schema not in ALLOWED_SCHEMAS:
        raise HTTPException(400, "unknown schema")
    if not files:
        raise HTTPException(400, "no files")

    created_here = False
    if session_id:
        sess = _real_sessions.get(session_id)
        if not sess or not sess["dir"].exists():
            raise HTTPException(404, "no such real-pile session")
        sess_dir = sess["dir"]
        existing = sum(1 for _ in sess_dir.iterdir()
                       if _.is_file() and _.name not in {"extraction.json"})
        if existing + len(files) > MAX_UPLOAD_FILES:
            raise HTTPException(400, f"pile limit {MAX_UPLOAD_FILES} files "
                                     f"(currently has {existing})")
    else:
        if len(files) > MAX_UPLOAD_FILES:
            raise HTTPException(400, f"1-{MAX_UPLOAD_FILES} files required")
        session_id = uuid.uuid4().hex[:12]
        sess_dir = REAL / session_id
        sess_dir.mkdir(parents=True, exist_ok=True)
        created_here = True

    # Buffer + validate everything before touching disk. If any check
    # fails we can bail cleanly, and if we created the session dir in
    # this call, tear it down so we don't leak an empty session.
    try:
        buffered: list[tuple[str, bytes]] = []
        for uf in files:
            name = Path(uf.filename or "doc").name  # strip any path parts
            suffix = Path(name).suffix.lower()
            if suffix not in ALLOWED_UPLOAD_SUFFIXES:
                raise HTTPException(400, f"unsupported file type: {suffix}")
            content = await uf.read()
            if len(content) > MAX_UPLOAD_BYTES:
                raise HTTPException(413, f"{name} is {len(content) / 1024 / 1024:.1f} MB, over the "
                                         f"{MAX_UPLOAD_BYTES // 1024 // 1024} MB per-file limit")
            buffered.append((name, content))
        # All good — flush to disk in one pass.
        for name, content in buffered:
            (sess_dir / name).write_bytes(content)
    except HTTPException:
        if created_here:
            shutil.rmtree(sess_dir, ignore_errors=True)
            _real_sessions.pop(session_id, None)
        raise

    _real_sessions[session_id] = {"dir": sess_dir, "schema": schema}
    # Report exactly what got persisted so the client can track the
    # added batch and roll it back precisely on cancel.
    added_names = [name for name, _ in buffered]
    return {"session_id": session_id,
            "files": added_names,
            "added": added_names}


class RollbackBatchBody(BaseModel):
    session_id: str
    files: list[str]


@app.post("/api/real/rollback_batch")
def real_rollback_batch(body: RollbackBatchBody):
    """Delete a specific set of files from an existing session without
    tearing the whole session down. Wired to the Cancel button when
    the in-flight run is an 'Add more docs' batch — cancelling one
    batch should not vaporise the pile that was already there."""
    sess = _real_session(body.session_id)
    sess_dir = sess["dir"]
    removed = []
    for name in body.files:
        # Strip path parts; only touch files inside the session dir.
        safe = Path(name).name
        if not safe or safe.startswith('.') or safe in {"extraction.json"}:
            continue
        target = sess_dir / safe
        if target.exists() and target.is_file():
            target.unlink()
            removed.append(safe)
    # Invalidate any cached router since the pile just changed.
    _routers.pop(f"real:{body.session_id}", None)
    return {"ok": True, "removed": removed}


class RealSchemaBody(BaseModel):
    session_id: str
    schema: str


@app.post("/api/real/set_schema")
def real_set_schema(body: RealSchemaBody):
    if body.schema not in ALLOWED_SCHEMAS:
        raise HTTPException(400, "unknown schema")
    sess = _real_session(body.session_id)
    sess["schema"] = body.schema
    _routers.pop(f"real:{body.session_id}", None)
    return {"ok": True}


class RealRunBody(BaseModel):
    session_id: str
    full_local: bool | None = None


@app.post("/api/real/run")
def real_run(body: RealRunBody):
    """Extract → redact → reconcile the uploaded pile. Extraction hits the
    same Gemma endpoint as the demo piles; if VLLM_URL isn't reachable,
    the pipeline degrades to text-only ingest (PDFs may return empty)."""
    sess = _real_session(body.session_id)
    schema_path = ROOT / "paperflow" / "schemas" / f"{sess['schema']}.yaml"
    full_local = (body.full_local if body.full_local is not None
                  else "FIREWORKS_API_KEY" not in os.environ)

    # extract via live Gemma if reachable; cache the JSON for the pile view.
    # If extraction.json is already present (samples pre-fill it from
    # outputs/extraction_<folder>.json so demos work with Gemma asleep),
    # skip the live call and use the cached artefact.
    extraction_path = sess["dir"] / "extraction.json"
    if extraction_path.exists() and extraction_path.stat().st_size > 0:
        pass  # cached: samples ship with a pre-parsed extraction
    else:
        try:
            result = asyncio.run(extract_pile(sess["dir"], schema_path))
            extraction_path.write_text(json.dumps(result.to_dict(), indent=1))
        except Exception as e:  # noqa: BLE001 - remote unavailable: keep text
            extraction_path.write_text(json.dumps(
                {"pile": sess["dir"].name, "docs": [], "error": str(e)}))
            raise HTTPException(502, f"extractor unavailable: {type(e).__name__}. "
                                     f"Start a Gemma/vLLM endpoint and set "
                                     f"VLLM_URL, or try text-only files.")

    run_pile(sess["dir"], schema_path, extraction_path, full_local,
             out_root=sess["dir"] / "runs")
    # relocate to a stable location the router can open
    src = sess["dir"] / "runs" / f"run_{sess['dir'].name}"
    dst = sess["dir"] / "run"
    if dst.exists():
        shutil.rmtree(dst)
    shutil.move(str(src), str(dst))
    _routers.pop(f"real:{body.session_id}", None)
    return {"ok": True, "session_id": body.session_id}


@app.get("/api/real/pile/{session_id}")
def real_pile(session_id: str):
    sess = _real_session(session_id)
    run_dir = sess["dir"] / "run"
    extraction = sess["dir"] / "extraction.json"
    if not (run_dir / "run_output.json").exists() or not extraction.exists():
        raise HTTPException(409, "run the real pile first")
    schema_path = ROOT / "paperflow" / "schemas" / f"{sess['schema']}.yaml"
    view = build_pile_view(run_dir, extraction, schema_path)
    view["real"] = True
    return view


@app.get("/api/real/doc/{session_id}/{filename}")
def real_doc(session_id: str, filename: str):
    sess = _real_session(session_id)
    if "/" in filename or ".." in filename:
        raise HTTPException(404, "not found")
    path = sess["dir"] / filename
    if not path.exists():
        raise HTTPException(404, "not found")
    mime = {"pdf": "application/pdf", "txt": "text/plain",
            "md": "text/markdown",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }.get(path.suffix.lstrip("."), "application/octet-stream")
    return FileResponse(path, media_type=mime)


@app.get("/api/real/xlsx_preview/{session_id}/{filename}")
def real_xlsx_preview(session_id: str, filename: str, n: int = 24):
    """Same shape as /api/xlsx_preview but scoped to a Real-pile session.
    Uploaded XLSX files never got a preview because the client's xlsx
    branch only had a demo-pile URL to hit — this fills the gap."""
    sess = _real_session(session_id)
    if "/" in filename or ".." in filename:
        raise HTTPException(404, "not found")
    path = sess["dir"] / filename
    if not path.exists() or path.suffix != ".xlsx":
        raise HTTPException(404, "not found")
    return _read_xlsx_preview(path, n)


@app.post("/api/real/reset/{session_id}")
def real_reset(session_id: str):
    sess = _real_sessions.pop(session_id, None)
    if sess:
        shutil.rmtree(sess["dir"], ignore_errors=True)
    _routers.pop(f"real:{session_id}", None)
    return {"ok": True}
