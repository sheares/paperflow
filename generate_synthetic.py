#!/usr/bin/env python3
"""
Generate synthetic paperflow demo data.

Three piles under `synthetic/`:
- kyc_onboarding/    KYC forms, proof of address, NRIC scan, email declaration
- partner_collation/ Registration sheets, MoU, business cards, email footers
- patient_intake/    Intake forms, referral letters, lab requisition, insurance

Each PDF is composed from the same set of primitives (letterhead with a
document-reference meta bar, section bands, boxed data panels, ruled tables,
signature grids, barcodes, rubber stamps and a legal footer) so the pages
read like something someone would actually receive, not a stripped-down
mock-up. Every planted VALUE stays identical to keep the scorer's
ground-truth stable across regenerations.

`synthetic/samples/` also contains scanned-image documents (PNG/JPEG) for
the Real-pile upload demo.

Install: pip install reportlab openpyxl pillow
Run:     python3 generate_synthetic.py
"""
import json
import random
import shutil
import textwrap
from pathlib import Path

from PIL import Image, ImageDraw, ImageFilter, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font as XLFont, PatternFill, Side
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas


HERE = Path(__file__).parent
OUT = HERE / 'synthetic'

# ---- palette ------------------------------------------------------------
INK = HexColor('#111111')
INK_SOFT = HexColor('#2A2A2A')
MUTED = HexColor('#555555')
FAINT = HexColor('#8A8A8A')
HAIR = HexColor('#D0D0D0')
BG_SOFT = HexColor('#F4F5F7')
BG_BAND = HexColor('#EEF1F5')

PAGE_W, PAGE_H = A4
MARGIN = 18 * mm


# ============================================================
# LOW-LEVEL PRIMITIVES
# ============================================================

def _wordmark(c, x, y, text, brand, size=14):
    """Small square mark with a knocked-out disc + wordmark next to it.
    Reads as a plausible brand element without cloning any real logo."""
    c.setFillColor(brand)
    c.rect(x, y - 5.5 * mm, 7 * mm, 7 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor('#FFFFFF'))
    c.circle(x + 4.7 * mm, y - 2 * mm, 1.4 * mm, fill=1, stroke=0)
    c.setFillColor(brand)
    c.setFont('Helvetica-Bold', size)
    c.drawString(x + 9.5 * mm, y - 1 * mm, text)


def _rule(c, x1, y, x2, colour=HAIR, w=0.5):
    c.setStrokeColor(colour)
    c.setLineWidth(w)
    c.line(x1, y, x2, y)


def _accent(c, x1, y, x2, colour, w=1.6):
    c.setStrokeColor(colour)
    c.setLineWidth(w)
    c.line(x1, y, x2, y)


def _double_rule(c, x1, y, x2, colour, gap=1.2 * mm):
    """Two parallel rules — the visual signature of formal contracts and
    government headers."""
    c.setStrokeColor(colour); c.setLineWidth(1.2)
    c.line(x1, y, x2, y)
    c.setLineWidth(0.5)
    c.line(x1, y - gap, x2, y - gap)


def _letterhead(c, brand, name, tagline, address, contact, meta=None):
    """Full letterhead. Left: mark + wordmark + tagline + address block.
    Right: right-aligned meta lines (Ref, Date, Page, Series...).
    Returns the y position where body content should start."""
    x_left = MARGIN
    x_right = PAGE_W - MARGIN
    y = PAGE_H - MARGIN - 2 * mm

    _wordmark(c, x_left, y, name, brand)
    c.setFillColor(MUTED); c.setFont('Helvetica', 8.5)
    c.drawString(x_left + 9.5 * mm, y - 5.5 * mm, tagline)

    # left address block, small
    c.setFillColor(MUTED); c.setFont('Helvetica', 7.8)
    c.drawString(x_left, y - 12 * mm, address)
    c.drawString(x_left, y - 15.4 * mm, contact)

    # right meta lines (each a (label, value) pair rendered as
    # "Label   value" right-aligned)
    if meta:
        c.setFont('Helvetica', 8)
        my = y
        for label, value in meta:
            c.setFillColor(FAINT)
            c.drawRightString(x_right - 32 * mm, my - 1 * mm, label)
            c.setFillColor(INK); c.setFont('Helvetica-Bold', 8.6)
            c.drawRightString(x_right, my - 1 * mm, str(value))
            c.setFont('Helvetica', 8)
            my -= 3.6 * mm

    _accent(c, x_left, y - 19 * mm, x_right, brand, w=1.6)
    return y - 25 * mm


def _band_header(c, brand, name, tagline, address, contact, meta=None):
    """Alternative header: full-width coloured band with the wordmark in
    white on the brand, address on the right in light text. Used by
    clinical / statement / government documents where the letterhead is
    the whole top strip."""
    x_left = MARGIN
    x_right = PAGE_W - MARGIN
    band_h = 30 * mm
    y_top = PAGE_H - MARGIN + 5 * mm
    # brand-coloured band across the full width
    c.setFillColor(brand)
    c.rect(0, y_top - band_h, PAGE_W, band_h, fill=1, stroke=0)
    # brand mark on band: white square with brand-coloured knock-out
    c.setFillColor(HexColor('#FFFFFF'))
    c.rect(x_left, y_top - band_h + 12 * mm, 9 * mm, 9 * mm,
           fill=1, stroke=0)
    c.setFillColor(brand)
    c.circle(x_left + 6.5 * mm, y_top - band_h + 15 * mm, 2 * mm,
             fill=1, stroke=0)
    # wordmark
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(x_left + 12.5 * mm, y_top - band_h + 15 * mm, name)
    c.setFont('Helvetica', 8.5)
    c.setFillColor(HexColor('#EEEEEE'))
    c.drawString(x_left + 12.5 * mm, y_top - band_h + 10 * mm, tagline)
    # address + contact right-side, light
    c.setFillColor(HexColor('#F1F1F1'))
    c.setFont('Helvetica', 7.8)
    c.drawRightString(x_right, y_top - band_h + 18 * mm, address)
    c.drawRightString(x_right, y_top - band_h + 14 * mm, contact)
    # meta line underneath (label · value pairs separated by pipes),
    # bottom of the band
    if meta:
        parts = []
        for label, value in meta:
            parts.append(f'{label} {value}')
        c.setFillColor(HexColor('#F1F1F1'))
        c.setFont('Helvetica', 7.5)
        c.drawRightString(x_right, y_top - band_h + 4 * mm,
                          '  ·  '.join(parts))
    return y_top - band_h - 6 * mm


def _memo_header(c, brand, sender, address, contact, date, ref):
    """Third header style: classic memo/letter — centred sender at the
    very top, thin double rule under it, date + ref right-aligned. No
    coloured band. Used for referral letters and MoUs where the doc
    reads more like paper correspondence."""
    x_left = MARGIN
    x_right = PAGE_W - MARGIN
    y = PAGE_H - MARGIN - 2 * mm
    c.setFillColor(brand); c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(PAGE_W / 2, y, sender)
    c.setFillColor(MUTED); c.setFont('Helvetica', 8.5)
    c.drawCentredString(PAGE_W / 2, y - 5.5 * mm, address)
    c.drawCentredString(PAGE_W / 2, y - 9 * mm, contact)
    _double_rule(c, x_left, y - 13 * mm, x_right, brand)
    # date + ref right-aligned under the double rule
    c.setFillColor(MUTED); c.setFont('Helvetica', 8)
    c.drawRightString(x_right, y - 18 * mm, f'Date {date}')
    c.drawRightString(x_right, y - 21.5 * mm, f'Ref  {ref}')
    return y - 27 * mm


def _title(c, x, y, text, sub=None):
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 15)
    c.drawString(x, y, text)
    if sub:
        c.setFillColor(MUTED); c.setFont('Helvetica', 9)
        c.drawString(x, y - 5.6 * mm, sub)
        return y - 12 * mm
    return y - 8 * mm


def _numbered_section(c, x, y, num, title, brand):
    """Alternative to _section_band: numbered clause header used by
    contracts and formal declarations. No shaded band — just a big
    numbered lead-in with a thin brand underline."""
    c.setFillColor(brand); c.setFont('Helvetica-Bold', 13)
    c.drawString(x, y, str(num))
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 11)
    c.drawString(x + 9 * mm, y, title)
    _rule(c, x, y - 2 * mm, x + 60 * mm, colour=brand, w=0.8)
    return y - 8 * mm


def _amount_due_block(c, x, y, w, brand, amount, due_date, account_no):
    """Big amount-due panel used at the top-right of statements and
    invoices. Bright brand-coloured strip with the number in a large
    face; below, an account number line."""
    h = 32 * mm
    c.setFillColor(brand)
    c.rect(x, y - h, w, h, fill=1, stroke=0)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont('Helvetica-Bold', 8)
    c.drawString(x + 4 * mm, y - 6 * mm, 'AMOUNT DUE')
    c.setFont('Helvetica-Bold', 24)
    c.drawString(x + 4 * mm, y - 16 * mm, amount)
    c.setFont('Helvetica', 8)
    c.drawString(x + 4 * mm, y - 22 * mm, f'Due by {due_date}')
    c.drawString(x + 4 * mm, y - 26 * mm, f'Account {account_no}')
    return y - h - 4 * mm


def _section_band(c, y, letter, title, brand):
    """Sectioned form band: coloured letter box + shaded band with the
    section title in caps. This is what real bank/hospital forms use to
    split personal particulars from declarations from signatures."""
    x_left = MARGIN
    x_right = PAGE_W - MARGIN
    c.setFillColor(BG_BAND)
    c.rect(x_left, y - 6 * mm, x_right - x_left, 6 * mm, fill=1, stroke=0)
    c.setFillColor(brand)
    c.rect(x_left, y - 6 * mm, 6 * mm, 6 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont('Helvetica-Bold', 10)
    c.drawCentredString(x_left + 3 * mm, y - 4.2 * mm, letter)
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 9.5)
    c.drawString(x_left + 9 * mm, y - 4.2 * mm, title.upper())
    return y - 10 * mm


def _prose(c, x, y, text, width_chars=95, size=9, leading=4.4 * mm,
           font='Helvetica', colour=INK_SOFT):
    c.setFillColor(colour); c.setFont(font, size)
    for para in text.split('\n\n'):
        for line in textwrap.wrap(para.strip(), width=width_chars):
            c.drawString(x, y, line)
            y -= leading
        y -= leading * 0.4
    return y


def _fields(c, x, y, rows, label_w=48 * mm, gap=5.5 * mm, label_size=9,
            value_size=10):
    """Vertical label: value stack. Labels keep the colon because
    pdftotext -layout preserves the label-value structure the redactor's
    scan pass relies on."""
    for row in rows:
        if row == '':
            y -= gap * 0.5
            continue
        if isinstance(row, tuple):
            label, value = row
            if label:
                c.setFillColor(MUTED); c.setFont('Helvetica-Bold', label_size)
                c.drawString(x, y, f'{label}:')
            c.setFillColor(INK); c.setFont('Helvetica', value_size)
            c.drawString(x + label_w, y, str(value))
        else:
            c.setFillColor(INK); c.setFont('Helvetica', value_size)
            c.drawString(x, y, str(row))
        y -= gap
    return y


def _zebra_fields(c, x, y, rows, w=None, label_w=48 * mm, row_h=6 * mm,
                  band=BG_SOFT, label_size=8.8, value_size=10):
    """Zebra-striped label:value rows — the second row shaded, third
    clear, etc. Very common on clinical intake forms and enrolment
    sheets because it keeps the eye tracking across the page."""
    if w is None:
        w = PAGE_W - 2 * MARGIN
    for i, row in enumerate(rows):
        if row == '':
            y -= row_h * 0.4
            continue
        if i % 2 == 0:
            c.setFillColor(band)
            c.rect(x, y - row_h + 1 * mm, w, row_h, fill=1, stroke=0)
        if isinstance(row, tuple):
            label, value = row
            if label:
                c.setFillColor(MUTED); c.setFont('Helvetica-Bold', label_size)
                c.drawString(x + 3 * mm, y - 2.8 * mm, f'{label}:')
            c.setFillColor(INK); c.setFont('Helvetica', value_size)
            c.drawString(x + 3 * mm + label_w, y - 2.8 * mm, str(value))
        y -= row_h
    return y - 2 * mm


def _grid_fields(c, x, y, rows, col_w=85 * mm, label_w=32 * mm,
                 gap=6 * mm, cols=2):
    """Two-column label: value grid for dense demographic panels. Rows
    are laid out row-major: rows[0] top-left, rows[1] top-right, rows[2]
    second row left, and so on."""
    for i, row in enumerate(rows):
        col = i % cols
        line = i // cols
        cx = x + col * col_w
        cy = y - line * gap
        if isinstance(row, tuple):
            label, value = row
            c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 8.5)
            c.drawString(cx, cy, f'{label}:')
            c.setFillColor(INK); c.setFont('Helvetica', 9.5)
            c.drawString(cx + label_w, cy, str(value))
    n_lines = (len(rows) + cols - 1) // cols
    return y - n_lines * gap - 2 * mm


def _boxed_panel(c, x, y, w, title, rows, brand, label_w=42 * mm,
                 row_gap=5.5 * mm):
    """A framed panel with a shaded title bar and label: value rows.
    Used for account summaries, patient particulars, insurance details."""
    n = sum(1 for r in rows if r != '')
    h = 8 * mm + n * row_gap + 4 * mm
    # title bar
    c.setFillColor(brand)
    c.rect(x, y - 6 * mm, w, 6 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont('Helvetica-Bold', 9)
    c.drawString(x + 3 * mm, y - 4.2 * mm, title.upper())
    # body outline
    c.setStrokeColor(HAIR); c.setLineWidth(0.6)
    c.setFillColor(HexColor('#FFFFFF'))
    c.rect(x, y - h, w, h - 6 * mm, fill=1, stroke=1)
    # rows
    _fields(c, x + 3 * mm, y - 10 * mm, rows,
            label_w=label_w, gap=row_gap, label_size=8.5, value_size=9.5)
    return y - h - 2 * mm


def _two_boxes(c, y, left_title, left_rows, right_title, right_rows,
               brand, gap=4 * mm):
    """Side-by-side boxed panels. Used for patient particulars + referring
    clinician on lab requisitions."""
    x_left = MARGIN
    x_right = PAGE_W - MARGIN
    box_w = (x_right - x_left - gap) / 2
    y_after_left = _boxed_panel(c, x_left, y, box_w, left_title, left_rows,
                                brand, label_w=30 * mm)
    y_after_right = _boxed_panel(c, x_left + box_w + gap, y, box_w,
                                 right_title, right_rows, brand,
                                 label_w=30 * mm)
    return min(y_after_left, y_after_right)


def _ruled_table(c, x, y, headers, rows, col_widths, header_bg=BG_BAND,
                 row_h=5.5 * mm, header_h=6 * mm, font_size=8.5):
    """Ruled table with a shaded header row and light gridlines. Every
    row draws its own bottom rule so pdftotext -layout keeps the columns
    aligned."""
    total_w = sum(col_widths)
    # header band
    c.setFillColor(header_bg)
    c.rect(x, y - header_h, total_w, header_h, fill=1, stroke=0)
    c.setFillColor(INK); c.setFont('Helvetica-Bold', font_size)
    cx = x
    for label, w in zip(headers, col_widths):
        c.drawString(cx + 1.4 * mm, y - header_h + 1.8 * mm, label)
        cx += w
    _rule(c, x, y - header_h, x + total_w)
    ry = y - header_h
    c.setFont('Helvetica', font_size)
    for row in rows:
        cx = x
        for cell, w in zip(row, col_widths):
            c.setFillColor(INK)
            c.drawString(cx + 1.4 * mm, ry - row_h + 1.8 * mm, str(cell))
            cx += w
        ry -= row_h
        _rule(c, x, ry, x + total_w)
    # outline
    c.setStrokeColor(HAIR); c.setLineWidth(0.6)
    c.rect(x, ry, total_w, y - ry, fill=0, stroke=1)
    # vertical dividers
    cx = x
    for w in col_widths[:-1]:
        cx += w
        c.setStrokeColor(HAIR); c.setLineWidth(0.4)
        c.line(cx, ry, cx, y)
    return ry - 3 * mm


def _tick_row(c, x, y, question, options=('Yes', 'No'), selected='No'):
    """Yes / No / N/A tick-box row. Puts a small ☒/☐ pair after the
    question."""
    c.setFillColor(INK); c.setFont('Helvetica', 9)
    c.drawString(x, y, question)
    cx = x + 130 * mm
    for opt in options:
        # box
        c.setStrokeColor(INK); c.setLineWidth(0.6)
        c.rect(cx, y - 0.7 * mm, 3 * mm, 3 * mm, fill=0, stroke=1)
        if opt == selected:
            c.setFillColor(INK)
            c.setLineWidth(0.9)
            c.line(cx + 0.7 * mm, y + 0.5 * mm,
                   cx + 1.5 * mm, y - 0.2 * mm)
            c.line(cx + 1.5 * mm, y - 0.2 * mm,
                   cx + 2.6 * mm, y + 1.5 * mm)
        c.setFillColor(INK); c.setFont('Helvetica', 8.5)
        c.drawString(cx + 4 * mm, y, opt)
        cx += 15 * mm
    return y - 5.2 * mm


def _barcode(c, x, y, w, h, code, label_size=6.5):
    """Fake Code 128 vertical bars. Deterministic per code string so
    the same doc always renders the same barcode."""
    r = random.Random(hash(code) & 0xFFFF)
    c.setFillColor(INK)
    cx = x
    while cx < x + w - 0.5:
        bw = r.choice([0.3, 0.6, 0.8, 1.1, 1.4])
        gap = r.choice([0.3, 0.5, 0.7])
        c.rect(cx, y, bw, h, fill=1, stroke=0)
        cx += bw + gap
    c.setFont('Helvetica', label_size); c.setFillColor(MUTED)
    c.drawString(x, y - 2.4 * mm, code)


def _stamp(c, cx, cy, main, sub, brand=HexColor('#8B2018'), r=15 * mm):
    """Round rubber-stamp overlay. Draws a double-ring circle with the
    stamp text arranged main-on-top, sub-on-bottom."""
    c.saveState()
    c.setStrokeColor(brand); c.setLineWidth(1.4)
    c.circle(cx, cy, r, fill=0, stroke=1)
    c.circle(cx, cy, r - 1.6 * mm, fill=0, stroke=1)
    c.setFillColor(brand); c.setFont('Helvetica-Bold', 9)
    c.drawCentredString(cx, cy + 1.2 * mm, main)
    c.setFont('Helvetica', 7)
    c.drawCentredString(cx, cy - 3 * mm, sub)
    c.restoreState()


def _signature_grid(c, x, y, w, entries, gap=6 * mm):
    """Signature grid. Each entry is a dict with printed_name, role,
    signature_hint, id_label, id_value, date. Renders label rows +
    signed name + underlined date/id lines."""
    col_w = (w - gap * (len(entries) - 1)) / len(entries)
    cy = y
    for i, e in enumerate(entries):
        cx = x + i * (col_w + gap)
        # signature line
        c.setStrokeColor(INK); c.setLineWidth(0.6)
        c.line(cx, cy - 12 * mm, cx + col_w - 8 * mm, cy - 12 * mm)
        c.setFillColor(MUTED); c.setFont('Helvetica', 7.5)
        c.drawString(cx, cy - 14.5 * mm, 'Signature')
        # printed name + role
        c.setFillColor(INK); c.setFont('Helvetica-Bold', 9.5)
        c.drawString(cx, cy - 20 * mm, e['printed_name'])
        c.setFillColor(MUTED); c.setFont('Helvetica', 8)
        c.drawString(cx, cy - 23.5 * mm, e['role'])
        # id (NRIC / MCR / registration) + date
        c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 7.5)
        c.drawString(cx, cy - 28 * mm, f"{e['id_label']}:")
        c.setFillColor(INK); c.setFont('Helvetica', 8.5)
        c.drawString(cx + 20 * mm, cy - 28 * mm, e['id_value'])
        c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 7.5)
        c.drawString(cx, cy - 32 * mm, 'Date:')
        c.setFillColor(INK); c.setFont('Helvetica', 8.5)
        c.drawString(cx + 20 * mm, cy - 32 * mm, e.get('date', ''))
    return cy - 36 * mm


def _footer_legal(c, lines):
    c.setFillColor(FAINT); c.setFont('Helvetica', 6.8)
    y = 12 * mm
    for line in lines:
        c.drawString(MARGIN, y, line)
        y -= 2.6 * mm


def _watermark(c):
    c.setFillColor(FAINT); c.setFont('Helvetica-Oblique', 6.5)
    c.drawString(MARGIN, 6 * mm,
                 'Synthetic document · paperflow demo · not a real record')


def _page_num(c, page, of):
    c.setFillColor(MUTED); c.setFont('Helvetica', 7.5)
    c.drawRightString(PAGE_W - MARGIN, 6 * mm, f'Page {page} of {of}')


# ============================================================
# TXT / XLSX / JSON HELPERS
# ============================================================

def write_txt(path: Path, lines):
    out = []
    for line in lines:
        if isinstance(line, tuple):
            label, value = line
            out.append(f'{label}: {value}' if label else str(value))
        else:
            out.append(str(line))
    path.write_text('\n'.join(out) + '\n')


def write_xlsx(path: Path, sheet_name: str, rows, extra_rows=None,
               title=None, subtitle=None, note=None,
               title_bg='1F3B73'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    thin = Side(border_style='thin', color='CCCCCC')
    header_fill = PatternFill('solid', fgColor='F0F2F5')
    title_fill = PatternFill('solid', fgColor=title_bg)

    r = 1
    if title:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = XLFont(bold=True, color='FFFFFF', size=13)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[r].height = 26
        r += 1
    if subtitle:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=subtitle)
        cell.font = XLFont(italic=True, color='585858', size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        r += 1
    r += 1
    for k, v in rows:
        key_cell = ws.cell(row=r, column=1, value=k)
        key_cell.font = XLFont(bold=True)
        key_cell.fill = header_fill
        key_cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        val_cell = ws.cell(row=r, column=2, value=v)
        val_cell.alignment = Alignment(vertical='center')
        val_cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.row_dimensions[r].height = 20
        r += 1
    if extra_rows:
        r += 1
        for k, v in extra_rows:
            ws.cell(row=r, column=1, value=k).font = XLFont(italic=True, color='585858')
            ws.cell(row=r, column=2, value=v).font = XLFont(color='585858')
            r += 1
    if note:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=note)
        cell.font = XLFont(italic=True, color='8A8A8A', size=9)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 54
    wb.save(str(path))


def write_json(path: Path, data: dict):
    path.write_text(json.dumps(data, indent=2))


# ============================================================
# BITMAP DOCUMENTS (samples/ folder for the Real-pile demo)
# ============================================================

def _load_font(pref_list, size):
    for name in pref_list:
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


BODY_FONT_CANDIDATES = [
    '/System/Library/Fonts/Supplemental/Arial.ttf',
    '/System/Library/Fonts/Helvetica.ttc',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
]
BOLD_FONT_CANDIDATES = [
    '/System/Library/Fonts/Supplemental/Arial Bold.ttf',
    '/Library/Fonts/Arial Bold.ttf',
    '/System/Library/Fonts/Helvetica.ttc',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
]


def _paper_texture(img: Image.Image):
    r = random.Random(42)
    px = img.load()
    w, h = img.size
    for _ in range(w * h // 32):
        x, y = r.randrange(w), r.randrange(h)
        v = r.randint(-10, 4)
        p = px[x, y]
        px[x, y] = tuple(max(0, min(255, c + v)) for c in p[:3])
    return img.filter(ImageFilter.GaussianBlur(0.3))


def _render_scan_jpeg(path: Path, title: str, rows: list[tuple[str, str]],
                      brand=(210, 25, 30), rotate=1.2,
                      size=(1200, 1650)):
    img = Image.new('RGB', size, (252, 250, 246))
    draw = ImageDraw.Draw(img)
    w, h = size

    title_font = _load_font(BOLD_FONT_CANDIDATES, 44)
    tag_font = _load_font(BODY_FONT_CANDIDATES, 18)
    lbl_font = _load_font(BOLD_FONT_CANDIDATES, 22)
    val_font = _load_font(BODY_FONT_CANDIDATES, 24)
    small_font = _load_font(BODY_FONT_CANDIDATES, 14)

    draw.rectangle([80, 80, 130, 130], fill=brand)
    draw.ellipse([100, 95, 122, 117], fill=(255, 255, 255))
    draw.text((150, 80), title.split('·')[0].strip(),
              font=title_font, fill=(20, 20, 20))
    draw.text((150, 140), title.split('·', 1)[1].strip() if '·' in title
              else '', font=tag_font, fill=(90, 90, 90))
    draw.line([(80, 180), (w - 80, 180)], fill=brand, width=3)

    y = 240
    for row in rows:
        if row == '':
            y += 20; continue
        if isinstance(row, tuple):
            label, value = row
            if label:
                draw.text((80, y), f'{label}:', font=lbl_font, fill=(70, 70, 70))
            draw.text((400, y), str(value), font=val_font, fill=(15, 15, 15))
        else:
            draw.text((80, y), str(row), font=val_font, fill=(15, 15, 15))
        y += 55

    y_stamp = h - 380
    r = random.Random(hash(path.name) & 0xFFFF)
    sx, sy = w - 320 + r.randint(-10, 10), y_stamp + r.randint(-5, 5)
    for offset in range(3):
        draw.ellipse(
            [sx + offset, sy + offset, sx + 220 - offset, sy + 160 - offset],
            outline=(120, 20, 20), width=3)
    draw.text((sx + 34, sy + 40), 'RECEIVED', font=lbl_font,
              fill=(120, 20, 20))
    draw.text((sx + 44, sy + 90), 'FILE COPY', font=small_font,
              fill=(120, 20, 20))

    draw.line([(80, h - 130), (w - 80, h - 130)], fill=(200, 200, 200), width=1)
    draw.text((80, h - 100), 'Synthetic document · paperflow demo · not a real record',
              font=small_font, fill=(140, 140, 140))

    img = _paper_texture(img)
    if rotate:
        img = img.rotate(rotate, resample=Image.BICUBIC, fillcolor=(252, 250, 246), expand=False)
    img.save(path, 'JPEG', quality=88)


def _render_id_card_png(path: Path, name: str, nric: str,
                        dob: str = '02 Nov 1994', size=(1000, 620)):
    img = Image.new('RGB', size, (215, 220, 210))
    draw = ImageDraw.Draw(img)
    w, h = size

    card = Image.new('RGB', (860, 500), (238, 232, 214))
    cdraw = ImageDraw.Draw(card)
    cdraw.rectangle([0, 0, 860, 90], fill=(170, 25, 25))
    header_font = _load_font(BOLD_FONT_CANDIDATES, 26)
    sub_font = _load_font(BODY_FONT_CANDIDATES, 16)
    lbl_font = _load_font(BOLD_FONT_CANDIDATES, 18)
    val_font = _load_font(BODY_FONT_CANDIDATES, 22)
    cdraw.text((30, 22), 'REPUBLIC OF SINGAPORE',
               font=header_font, fill=(255, 255, 255))
    cdraw.text((30, 55),
               'NATIONAL REGISTRATION IDENTITY CARD',
               font=sub_font, fill=(255, 240, 240))

    cdraw.rectangle([30, 120, 240, 380], fill=(180, 175, 165))
    cdraw.ellipse([80, 155, 190, 265], fill=(140, 135, 125))
    cdraw.rectangle([70, 275, 200, 380], fill=(150, 145, 135))

    cdraw.text((280, 130), 'Name', font=lbl_font, fill=(80, 60, 60))
    cdraw.text((280, 158), name, font=val_font, fill=(30, 25, 20))
    cdraw.text((280, 220), 'NRIC No.', font=lbl_font, fill=(80, 60, 60))
    cdraw.text((280, 248), nric, font=val_font, fill=(30, 25, 20))
    cdraw.text((280, 310), 'Date of Birth', font=lbl_font, fill=(80, 60, 60))
    cdraw.text((280, 338), dob, font=val_font, fill=(30, 25, 20))

    cdraw.line([(280, 420), (300, 415), (330, 430), (360, 410),
                (400, 425), (440, 415)], fill=(30, 20, 20), width=2)

    card = card.rotate(-1.2, resample=Image.BICUBIC, expand=True,
                       fillcolor=(215, 220, 210))
    px = (w - card.width) // 2
    py = (h - card.height) // 2
    img.paste(card, (px, py))
    for i in range(80):
        alpha = int(80 * (1 - i / 80))
        img.putalpha(255)
        overlay = Image.new('RGBA', img.size, (0, 0, 0, 0))
        odraw = ImageDraw.Draw(overlay)
        odraw.rectangle([i, i, w - i, h - i], outline=(0, 0, 0, alpha // 8))
        img = Image.alpha_composite(img.convert('RGBA'), overlay).convert('RGB')

    img = _paper_texture(img)
    img.save(path, 'PNG')


def _render_signed_letter_jpeg(path: Path, header: str, tagline: str,
                               brand, body_paragraphs, sign_off_name,
                               sign_off_title, size=(1240, 1650)):
    img = Image.new('RGB', size, (250, 249, 244))
    draw = ImageDraw.Draw(img)
    w, h = size

    title_font = _load_font(BOLD_FONT_CANDIDATES, 34)
    tag_font = _load_font(BODY_FONT_CANDIDATES, 16)
    body_font = _load_font(BODY_FONT_CANDIDATES, 22)
    sig_name_font = _load_font(BOLD_FONT_CANDIDATES, 24)
    small_font = _load_font(BODY_FONT_CANDIDATES, 14)

    draw.rectangle([90, 90, 140, 140], fill=brand)
    draw.polygon([(100, 100), (130, 100), (115, 130)], fill=(255, 255, 255))
    draw.text((160, 90), header, font=title_font, fill=(20, 20, 20))
    draw.text((160, 140), tagline, font=tag_font, fill=(90, 90, 90))
    draw.line([(90, 175), (w - 90, 175)], fill=brand, width=3)

    y = 240
    for para in body_paragraphs:
        for line in textwrap.wrap(para, width=72):
            draw.text((90, y), line, font=body_font, fill=(30, 30, 30))
            y += 32
        y += 18

    y += 20
    draw.text((90, y), 'Yours sincerely,',
              font=body_font, fill=(30, 30, 30))
    y += 40
    r = random.Random(hash(sign_off_name) & 0xFFFF)
    xs = 90
    for i in range(6):
        draw.line([(xs + i * 22 + r.randint(-4, 4),
                    y + r.randint(-4, 8)),
                   (xs + (i + 1) * 22 + r.randint(-4, 4),
                    y + r.randint(-12, 6))],
                  fill=(20, 25, 60), width=3)
    y += 40
    draw.text((90, y), sign_off_name, font=sig_name_font, fill=(20, 20, 20))
    y += 32
    draw.text((90, y), sign_off_title, font=tag_font, fill=(90, 90, 90))

    draw.line([(90, h - 130), (w - 90, h - 130)],
              fill=(200, 200, 200), width=1)
    draw.text((90, h - 100),
              'Synthetic document · paperflow demo · not a real record',
              font=small_font, fill=(140, 140, 140))

    img = _paper_texture(img).rotate(0.4, resample=Image.BICUBIC,
                                     fillcolor=(250, 249, 244), expand=False)
    img.save(path, 'JPEG', quality=90)


# ============================================================
# BRAND COLOURS (per doc issuer)
# ============================================================

DBS_BRAND = HexColor('#C4231F')
SP_BRAND = HexColor('#00A9CE')
NRIC_BRAND = HexColor('#8B1A1A')
HEART_BRAND = HexColor('#0072CE')
AIA_BRAND = HexColor('#D71920')
INNO_BRAND = HexColor('#6A1B9A')
ACME_BRAND = HexColor('#0B3D91')
BRIGHT_BRAND = HexColor('#2E7D32')
COBALT_BRAND = HexColor('#1F3B73')


# ============================================================
# KYC PILE
# ============================================================

def build_kyc():
    d = OUT / 'kyc_onboarding'
    d.mkdir(parents=True, exist_ok=True)

    # ---- 1. KYC form (DBS Private) — 2 pages, sectioned
    c = canvas.Canvas(str(d / 'kyc_form_hassan.pdf'), pagesize=A4)

    # PAGE 1
    kyc_meta = [
        ('Reference No.', 'KYC/2026/05/00218'),
        ('Case ID', 'CS-4471-2026'),
        ('Branch', 'Marina Bay Financial Centre'),
        ('RM Officer', 'Aloysius Tan (RM-2841)'),
        ('Received', '2026-05-22 · 10:14 SGT'),
    ]
    y = _letterhead(c, DBS_BRAND, 'DBS Private',
                    'Wealth Management · KYC/AML Onboarding',
                    '12 Marina Boulevard, DBS Asia Central @ Marina Bay Financial Centre Tower 3, Singapore 018982',
                    'Tel +65 6878 8888  ·  dbs.com.sg/private  ·  UEN 196800306E',
                    meta=kyc_meta)

    y = _title(c, MARGIN, y,
               'Know Your Customer (KYC) Declaration Form',
               'Individual Account · Wealth Onboarding Series 2026-B · Retain for compliance file')

    y = _prose(c, MARGIN, y,
               "I, the undersigned, declare that the information provided in this form is "
               "true, complete and accurate to the best of my knowledge. I understand that "
               "this declaration forms the basis of the Bank's assessment under the "
               "Monetary Authority of Singapore's Notice 626 (AML/CFT), and that any "
               "material misrepresentation may result in the account being suspended and "
               "referred to the Suspicious Transaction Reporting Office.")

    y = _section_band(c, y, 'A', 'Applicant particulars', DBS_BRAND)
    y = _grid_fields(c, MARGIN, y, [
        ('Doc ID', 'doc_001'),
        ('Nationality', 'Singaporean'),
        ('Full Name', 'Mohammed Farid bin Hassan'),
        ('Date of Birth', '1978-04-11'),
        ('National ID (NRIC)', 'K7741209'),
        ('Sex', 'Male'),
        ('Marital Status', 'Married'),
        ('Occupation', 'Managing Partner'),
    ])

    y = _section_band(c, y, 'B', 'Residential address', DBS_BRAND)
    y = _fields(c, MARGIN, y, [
        ('Residential Address', 'Blk 210 Bishan St 23 #11-04'),
        ('', 'Singapore 570210'),
        ('Length of Stay', '8 years, 3 months'),
        ('Previous Address', 'Blk 44 Toa Payoh Lor 5 #08-217, Singapore 310044'),
    ], gap=5 * mm)

    y = _section_band(c, y, 'C', 'Source of funds & employment', DBS_BRAND)
    y = _grid_fields(c, MARGIN, y, [
        ('Beneficial Owner', 'Self'),
        ('Employer', 'Farid & Partners LLP'),
        ('Source of Funds', ''),
        ('Monthly Income', 'SGD 32,000'),
        ('Declaration Date', '2026-05-22'),
        ('Employment Since', '2011-08-01'),
    ])

    _footer_legal(c, [
        'DBS Bank Ltd  ·  UEN 196800306E  ·  Registered office: 12 Marina Boulevard, DBS Asia Central Tower 3, Singapore 018982',
        'This form is strictly confidential and intended solely for the named applicant and DBS Wealth Onboarding compliance staff.',
        'Personal data is handled in accordance with the DBS Group Privacy Policy (dbs.com.sg/privacy).',
    ])
    _watermark(c); _page_num(c, 1, 2); c.showPage()

    # PAGE 2 — declarations + signature grid
    y = _letterhead(c, DBS_BRAND, 'DBS Private',
                    'Wealth Management · KYC/AML Onboarding',
                    '12 Marina Boulevard, DBS Asia Central @ Marina Bay Financial Centre Tower 3, Singapore 018982',
                    'Tel +65 6878 8888  ·  dbs.com.sg/private  ·  UEN 196800306E',
                    meta=kyc_meta)
    y = _title(c, MARGIN, y, 'Declarations & signature',
               'Section D-F · continued from previous page')

    y = _section_band(c, y, 'D', 'Politically-exposed person declaration', DBS_BRAND)
    y = _tick_row(c, MARGIN, y - 1 * mm,
                  '1. Are you a Politically Exposed Person (PEP)?',
                  selected='No')
    y = _tick_row(c, MARGIN, y,
                  '2. Are any immediate family members PEPs?',
                  selected='No')
    y = _tick_row(c, MARGIN, y,
                  '3. Are you a close associate of a PEP?',
                  selected='No')
    y -= 2 * mm

    y = _section_band(c, y, 'E', 'Declaration', DBS_BRAND)
    y = _prose(c, MARGIN, y,
               "1. I confirm that the funds referenced in Section C are derived from "
               "lawful sources.\n\n"
               "2. I consent to the Bank sharing my information with its authorised "
               "affiliates for the sole purpose of onboarding and ongoing due "
               "diligence.\n\n"
               "3. I understand that this declaration remains valid until superseded "
               "by a fresh declaration executed by me, and that I am obliged to "
               "update the Bank promptly if any material change occurs.")

    y = _section_band(c, y, 'F', 'Signatures', DBS_BRAND)
    y -= 6 * mm
    y = _signature_grid(c, MARGIN, y, PAGE_W - 2 * MARGIN, [
        {'printed_name': 'Mohammed Farid bin Hassan',
         'role': 'Applicant',
         'id_label': 'NRIC', 'id_value': 'K7741209',
         'date': '2026-05-22'},
        {'printed_name': 'Aloysius Tan',
         'role': 'Relationship Manager, DBS Private',
         'id_label': 'RM ID', 'id_value': 'RM-2841',
         'date': '2026-05-22'},
    ])

    _footer_legal(c, [
        'DBS Bank Ltd  ·  UEN 196800306E  ·  Registered office: 12 Marina Boulevard, DBS Asia Central Tower 3, Singapore 018982',
        'For internal use only — this page is part of a two-page KYC declaration and must not be circulated separately.',
    ])
    _watermark(c); _page_num(c, 2, 2); c.save()

    # ---- 2. Utility bill (SP Group) — band header + amount-due block
    #        (statement / invoice layout, distinct from KYC's sectioned form)
    c = canvas.Canvas(str(d / 'utility_bill_hassan.pdf'), pagesize=A4)
    y = _band_header(c, SP_BRAND, 'SP Group',
                     'Electricity, water and gas utilities · Regulated by EMA',
                     '2 Kallang Sector, Singapore 349277',
                     'Tel 1800 222 2333  ·  spgroup.com.sg  ·  UEN 199504676R',
                     meta=[('Invoice', 'SP-20260502-034172'),
                           ('Statement', '2026-05-02'),
                           ('Account', '03-124-5502-1')])

    # left: title + subtitle; right: big amount-due block
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 16)
    c.drawString(MARGIN, y - 4 * mm, 'Utility Bill · April 2026')
    c.setFillColor(MUTED); c.setFont('Helvetica', 9)
    c.drawString(MARGIN, y - 9 * mm,
                 'Household electricity, water and gas · 4-room HDB tariff')
    c.setFillColor(MUTED); c.setFont('Helvetica', 8)
    c.drawString(MARGIN, y - 14 * mm,
                 'Consumption in line with the seasonal average for a 4-room HDB flat.')
    _amount_due_block(c, PAGE_W - MARGIN - 60 * mm, y - 2 * mm,
                      60 * mm, SP_BRAND, 'SGD 84.30', '2026-05-22',
                      '03-124-5502-1')
    y -= 40 * mm

    # zebra-striped service details (statement style — no section band)
    y = _zebra_fields(c, MARGIN, y, [
        ('Doc ID', 'doc_002'),
        ('Account Holder', 'Farid Hassan'),
        ('Customer Ref (NRIC)', 'K7741209'),
        ('Service Address', 'Blk 88 Tampines Ave 4 #05-12, Singapore 521088'),
        ('Billing Period', '2026-04-01 to 2026-04-30'),
        ('Amount Due', 'SGD 84.30'),
    ], label_w=52 * mm)

    y -= 2 * mm
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 9.5)
    c.drawString(MARGIN, y, 'Charges breakdown')
    y -= 4 * mm
    y = _ruled_table(c, MARGIN, y,
                     headers=['Service', 'Previous', 'Current',
                              'Usage', 'Rate', 'Amount (SGD)'],
                     rows=[
                         ['Electricity (kWh)', '38,214', '38,528',
                          '314', '0.2145', '67.35'],
                         ['Water (m³)', '412.2', '419.8',
                          '7.6', '1.6540', '12.57'],
                         ['Gas (kWh)', '1,842', '1,868',
                          '26', '0.1621', '4.21'],
                         ['Rebate — U-Save', '', '', '', '', '-8.20'],
                         ['GST 9%', '', '', '', '', '7.55'],
                         ['Total due', '', '', '', '', '84.30'],
                     ],
                     col_widths=[46 * mm, 22 * mm, 22 * mm,
                                 20 * mm, 20 * mm, 24 * mm])

    y -= 2 * mm
    y = _prose(c, MARGIN, y,
               "Payment is due by 2026-05-22. Pay via AXS, GIRO, PayNow (to UEN "
               "199504676R), or at any SingPost branch. Please quote the account "
               "number above when contacting customer care. Late payments attract "
               "a reconnection fee of SGD 12.60.")

    # Payment stub with barcode
    y -= 4 * mm
    c.setDash(1, 3)
    _rule(c, MARGIN, y, PAGE_W - MARGIN, colour=MUTED, w=0.5)
    c.setDash([])
    y -= 5 * mm
    c.setFillColor(MUTED); c.setFont('Helvetica', 8)
    c.drawString(MARGIN, y, 'PAYMENT STUB · detach along the dashed line')
    y -= 5 * mm
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, f'Account 03-124-5502-1  ·  Amount SGD 84.30  ·  Due 2026-05-22')
    _barcode(c, MARGIN, y - 12 * mm, 80 * mm, 8 * mm,
             '0312455021||084.30||20260522')

    _footer_legal(c, [
        'SP Services Ltd  ·  UEN 199504676R  ·  2 Kallang Sector, Singapore 349277  ·  spgroup.com.sg',
        'This statement was generated from the electronic reading. Please retain for your records; a duplicate copy may incur a SGD 5 fee.',
    ])
    _watermark(c); c.save()

    # ---- 3. Yvonne's KYC declaration email (text file)
    write_txt(d / 'kyc_declaration_goh.txt', [
        'From: yvonne.goh@example.com',
        'To: kyc@bank.sg',
        'Cc: aloysius.tan@dbs.com',
        'Sent: 2026-05-24 14:12 SGT',
        'Subject: KYC Declaration — Y Goh (Ref KYC/2026/05/00219)',
        'X-Priority: Normal',
        '',
        ('Doc ID', 'doc_003'),
        '',
        'Hi team,',
        '',
        'Please find my KYC declaration below. I have reviewed the form and',
        'confirm the following details are accurate as at today. I will also',
        'attach a scan of my NRIC in a separate message so the two documents',
        'do not exceed the mailbox size cap.',
        '',
        ('Full Name', 'Yvonne Goh'),
        ('National ID (NRIC)', 'K3098551'),
        ('Residential Address', '8 Marina Boulevard #30-01, Singapore 018981'),
        ('Beneficial Owner', 'Self'),
        ('Source of Funds', 'Employment income'),
        '',
        "Note: I'll pop by the branch on Monday morning to sign the paper",
        'form. Kindly reach out if any of the above needs additional',
        'supporting documentation.',
        '',
        'Regards,',
        'Yvonne Goh',
        'Senior Product Manager, Fintech Practice',
        '8 Marina Boulevard #30-01, Singapore 018981',
        '+65 9123 4567',
        '',
        '--',
        'This email and any attachments are confidential. If you have received',
        'this in error please notify the sender and delete the message.',
    ])

    # ---- 4. NRIC scan wrapper — NRIC-shaped card rendered on A4
    c = canvas.Canvas(str(d / 'nric_scan_goh.pdf'), pagesize=A4)
    y = _letterhead(c, NRIC_BRAND, 'Republic of Singapore',
                    'National Registration Identity Card · Immigration & Checkpoints Authority',
                    '10 Kallang Road, ICA Building, Singapore 208718',
                    'Tel +65 6391 6100  ·  ica.gov.sg',
                    meta=[
                        ('Scan Date', '2026-05-24'),
                        ('Scanned By', 'DBS Wealth Onboarding'),
                        ('Case Ref.', 'KYC/2026/05/00219'),
                        ('Attached To', 'kyc_declaration_goh.txt'),
                    ])

    y = _title(c, MARGIN, y, 'ID Copy (Photostat)',
               "Scan submitted with Yvonne Goh's KYC declaration email · Section A verification")

    y = _prose(c, MARGIN, y,
               "The following details were transcribed from the National Registration "
               "Identity Card scan below. Small print on the physical card is often "
               "difficult to read cleanly on a photostat; any discrepancy against the "
               "written KYC declaration should be flagged for the compliance reviewer.")

    # NRIC-shaped card mockup (landscape rectangle)
    card_w = 130 * mm
    card_h = 82 * mm
    card_x = (PAGE_W - card_w) / 2
    card_y = y - card_h - 6 * mm
    c.setStrokeColor(HAIR); c.setLineWidth(0.6)
    c.setFillColor(HexColor('#EDE6D0'))
    c.roundRect(card_x, card_y, card_w, card_h, 4 * mm, fill=1, stroke=1)
    # header band
    c.setFillColor(NRIC_BRAND)
    c.rect(card_x, card_y + card_h - 16 * mm, card_w, 16 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont('Helvetica-Bold', 11)
    c.drawString(card_x + 5 * mm, card_y + card_h - 7 * mm,
                 'REPUBLIC OF SINGAPORE')
    c.setFont('Helvetica', 8)
    c.drawString(card_x + 5 * mm, card_y + card_h - 12 * mm,
                 'NATIONAL REGISTRATION IDENTITY CARD')
    # photo placeholder
    c.setFillColor(HexColor('#B7B0A0'))
    c.rect(card_x + 5 * mm, card_y + 6 * mm, 32 * mm, 44 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor('#8A8272'))
    c.circle(card_x + 21 * mm, card_y + 38 * mm, 6 * mm, fill=1, stroke=0)
    c.rect(card_x + 12 * mm, card_y + 12 * mm, 18 * mm, 18 * mm,
           fill=1, stroke=0)
    # data (transcription on the card face)
    tx = card_x + 42 * mm
    c.setFont('Helvetica-Bold', 8); c.setFillColor(HexColor('#655040'))
    c.drawString(tx, card_y + 46 * mm, 'Name')
    c.setFont('Helvetica-Bold', 12); c.setFillColor(INK)
    c.drawString(tx, card_y + 41 * mm, 'Yvonne Goh')
    c.setFont('Helvetica-Bold', 8); c.setFillColor(HexColor('#655040'))
    c.drawString(tx, card_y + 32 * mm, 'NRIC No.')
    c.setFont('Helvetica-Bold', 12); c.setFillColor(INK)
    c.drawString(tx, card_y + 27 * mm, 'K3098S51')
    c.setFont('Helvetica-Bold', 8); c.setFillColor(HexColor('#655040'))
    c.drawString(tx, card_y + 18 * mm, 'Date of Birth')
    c.setFont('Helvetica-Bold', 12); c.setFillColor(INK)
    c.drawString(tx, card_y + 13 * mm, '15 Sep 1988')
    # signature squiggle
    c.setStrokeColor(INK); c.setLineWidth(0.8)
    from reportlab.graphics.shapes import Path as _P  # noqa (placeholder)
    c.line(card_x + 42 * mm, card_y + 6 * mm,
           card_x + 55 * mm, card_y + 8 * mm)
    c.line(card_x + 55 * mm, card_y + 8 * mm,
           card_x + 65 * mm, card_y + 5 * mm)
    c.line(card_x + 65 * mm, card_y + 5 * mm,
           card_x + 82 * mm, card_y + 9 * mm)

    y = card_y - 6 * mm
    y = _boxed_panel(c, MARGIN, y, PAGE_W - 2 * MARGIN,
                     'Verification transcription',
                     [
                         ('Doc ID', 'doc_004'),
                         ('Name', 'Yvonne Goh'),
                         ('National ID (NRIC)', 'K3098S51'),
                         ('Date of Issue', '2018-11-22'),
                         ('Card Version', '3rd generation polycarbonate'),
                     ],
                     NRIC_BRAND, label_w=44 * mm)

    _footer_legal(c, [
        'This is a photocopy submitted for verification purposes only. The original card remains with the cardholder.',
        'Retention limited to the case file. Photocopies of the NRIC are subject to PDPA safeguards under the Advisory Guidelines for the Financial Sector.',
    ])
    _watermark(c); c.save()

    write_json(d / 'ground_truth.json', {
        'planted_conflicts': [
            {'field': 'Residential address', 'entity': 'Mohammed Farid bin Hassan',
             'per_doc': {'doc_001': 'Blk 210 Bishan St 23 #11-04 SG 570210',
                         'doc_002': 'Blk 88 Tampines Ave 4 #05-12 SG 521088'},
             'correct': 'Blk 210 Bishan St 23 #11-04 SG 570210'},
            {'field': 'National ID', 'entity': 'Yvonne Goh',
             'per_doc': {'doc_003': 'K3098551', 'doc_004': 'K3098S51'},
             'correct': 'K3098551', 'note': 'doc_004 OCR substituted 5 -> S'},
        ],
        'planted_gaps': [
            {'field': 'Source of funds', 'entity': 'Mohammed Farid bin Hassan',
             'reason': 'not captured on any document'},
            {'field': 'Declaration date', 'entity': 'Yvonne Goh',
             'reason': 'intake unsigned'},
        ],
        'alias_variations': [
            {'canonical': 'Mohammed Farid bin Hassan', 'aliases': ['Farid Hassan']},
        ],
        'sensitive_spans': [
            {'value': 'Mohammed Farid bin Hassan', 'type': 'PERSON'},
            {'value': 'Farid Hassan', 'type': 'PERSON'},
            {'value': 'Yvonne Goh', 'type': 'PERSON'},
            {'value': 'K7741209', 'type': 'NRIC'},
            {'value': 'K3098551', 'type': 'NRIC'},
            {'value': 'K3098S51', 'type': 'NRIC'},
            {'value': 'Blk 210 Bishan St 23 #11-04', 'type': 'ADDRESS'},
            {'value': 'Blk 88 Tampines Ave 4 #05-12', 'type': 'ADDRESS'},
            {'value': '8 Marina Boulevard #30-01', 'type': 'ADDRESS'},
            {'value': 'Singapore 570210', 'type': 'POSTCODE'},
            {'value': 'Singapore 521088', 'type': 'POSTCODE'},
            {'value': 'Singapore 018981', 'type': 'POSTCODE'},
        ],
    })


# ============================================================
# PARTNER PILE
# ============================================================

def build_partner():
    d = OUT / 'partner_collation'
    d.mkdir(parents=True, exist_ok=True)

    # ---- ACME registration (xlsx)
    write_xlsx(d / 'partner_registration_acme.xlsx', 'Partner Registration',
               title='PartnerConnect · Vendor Registration',
               subtitle='Registration Portal partners.event.sg  ·  Series 2026-B',
               title_bg='0B3D91',
               rows=[
                   ('Doc ID', 'doc_001'),
                   ('Organisation', 'Acme Robotics Pte Ltd'),
                   ('Contact Name', 'Lim Wei Jie'),
                   ('Email', 'weijie.lim@acmerobotics.sg'),
                   ('Phone', '+65 6123 4567'),
                   ('UEN', '201912345A'),
                   ('RSVP Status', 'Confirmed'),
               ],
               extra_rows=[
                   ('Submitted', '2026-05-14 09:41 SGT'),
                   ('Reviewed by', 'partners@event.sg'),
                   ('Portal Ref', 'PC-VR-2026-04-2211'),
               ],
               note='Submitted via the PartnerConnect portal. Please raise corrections through the vendor helpdesk (helpdesk@partners.event.sg).')

    # ---- ACME business card (transcription with a card mockup on the page)
    c = canvas.Canvas(str(d / 'acme_business_card.pdf'), pagesize=A4)
    y = _letterhead(c, ACME_BRAND, 'ACME Robotics',
                    'Industrial automation and integration services',
                    '12 Tuas Bay Walk, Singapore 638743',
                    'Tel +65 6789 1121  ·  acmerobotics.sg  ·  UEN 201912345A',
                    meta=[
                        ('Received', '2026-05-10 · SME Connect'),
                        ('Filed By', 'partners@event.sg'),
                        ('Attachment', 'card_scan.pdf'),
                    ])

    y = _title(c, MARGIN, y, 'Business card (scanned & transcribed)',
               'Received at the SME Connect networking session, 10 May 2026 · Vendor collation')

    y = _prose(c, MARGIN, y,
               "The following contact details were transcribed from a business card "
               "exchanged during the SME Connect networking session. The small print on "
               "the reverse of the card lists a different administrative email; both "
               "should be treated as valid contact points pending clarification.")

    # card mockup — small landscape rectangle on the page
    cw = 95 * mm; ch = 55 * mm
    cx = MARGIN + 5 * mm; cy = y - ch - 4 * mm
    c.setStrokeColor(HAIR); c.setLineWidth(0.5)
    c.setFillColor(HexColor('#FFFFFF'))
    c.roundRect(cx, cy, cw, ch, 2 * mm, fill=1, stroke=1)
    # brand strip
    c.setFillColor(ACME_BRAND)
    c.rect(cx, cy + ch - 8 * mm, cw, 8 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor('#FFFFFF')); c.setFont('Helvetica-Bold', 12)
    c.drawString(cx + 4 * mm, cy + ch - 5.5 * mm, 'ACME Robotics')
    c.setFont('Helvetica', 8)
    c.drawRightString(cx + cw - 4 * mm, cy + ch - 5.5 * mm,
                      'Industrial automation')
    # name + role
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 12)
    c.drawString(cx + 4 * mm, cy + ch - 16 * mm, 'Lim Wei Jie')
    c.setFillColor(MUTED); c.setFont('Helvetica', 8.5)
    c.drawString(cx + 4 * mm, cy + ch - 20 * mm,
                 'Head of Business Development')
    # contact
    c.setFillColor(INK); c.setFont('Helvetica', 8.5)
    c.drawString(cx + 4 * mm, cy + 16 * mm, 'wjlim@acme.com.sg')
    c.drawString(cx + 4 * mm, cy + 12 * mm, '+65 6123 4567')
    c.drawString(cx + 4 * mm, cy + 8 * mm, '+65 6789 1121 (direct)')
    c.drawString(cx + 4 * mm, cy + 4 * mm, '12 Tuas Bay Walk, Singapore 638743')
    # small print
    c.setFillColor(FAINT); c.setFont('Helvetica-Oblique', 6.5)
    c.drawString(cx + 4 * mm, cy + 0.5 * mm,
                 'UEN 201912345A · ACRA registered since 2019 · acmerobotics.sg')

    y = cy - 6 * mm
    y = _boxed_panel(c, MARGIN, y, PAGE_W - 2 * MARGIN,
                     'Transcription',
                     [
                         ('Doc ID', 'doc_002'),
                         ('Organisation', 'ACME Robotics'),
                         ('Contact', 'Lim Wei Jie'),
                         ('Role', 'Head of Business Development'),
                         ('Email', 'wjlim@acme.com.sg'),
                         ('Phone', '+65 6123 4567'),
                         ('Direct Line', '+65 6789 1121'),
                     ],
                     ACME_BRAND, label_w=42 * mm)

    _footer_legal(c, [
        'ACME Robotics Pte Ltd  ·  UEN 201912345A  ·  12 Tuas Bay Walk, Singapore 638743',
        'This transcription is the working copy retained by the event organiser and is not a formal declaration by the vendor.',
    ])
    _watermark(c); c.save()

    # ---- Brightpath MoU — memo/contract layout: centred sender, double
    #      rule, numbered clauses (not lettered section bands)
    c = canvas.Canvas(str(d / 'brightpath_mou.pdf'), pagesize=A4)
    y = _memo_header(c, BRIGHT_BRAND,
                     'Brightpath Learning LLP',
                     '15 Beach Road #04-08, Singapore 189677  ·  UEN T18LL0042K',
                     'Tel +65 9876 5432  ·  brightpath.edu.sg  ·  MoE approved enrichment provider',
                     date='2026-05-15',
                     ref='MOU/BRIGHTPATH/2026-11')

    c.setFillColor(INK); c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(PAGE_W / 2, y, 'MEMORANDUM OF UNDERSTANDING')
    c.setFillColor(MUTED); c.setFont('Helvetica-Oblique', 9)
    c.drawCentredString(PAGE_W / 2, y - 5.5 * mm,
                        'Community Partnership Programme · Series 2026-B')
    y -= 14 * mm

    y = _prose(c, MARGIN, y,
               "This Memorandum of Understanding (this \"MoU\") is entered into on "
               "15 May 2026 between the parties named below. It records the "
               "cooperative arrangements for the delivery of after-school "
               "enrichment programmes across three participating community centres "
               "during the second half of 2026, and is intended to be legally "
               "binding on the parties in respect of the confidentiality and "
               "payment terms only.")

    y = _numbered_section(c, MARGIN, y, '1.', 'Parties', BRIGHT_BRAND)
    y = _fields(c, MARGIN, y, [
        ('Doc ID', 'doc_003'),
        ('Party 1 (Provider)', 'Brightpath Learning LLP'),
        ('UEN', 'T18LL0042K'),
        ('Primary Contact', 'Nurul Aisyah'),
        ('Party 2 (Organiser)', 'PartnerConnect Events Pte Ltd'),
        ('Executed', '2026-05-15'),
    ], gap=5 * mm)

    y = _numbered_section(c, MARGIN, y, '2.', 'Term', BRIGHT_BRAND)
    y = _prose(c, MARGIN, y,
               "This MoU takes effect on 1 July 2026 and expires on 31 December "
               "2026, unless extended in writing signed by both parties.")

    y = _numbered_section(c, MARGIN, y, '3.', 'Responsibilities', BRIGHT_BRAND)
    y = _prose(c, MARGIN, y,
               "Brightpath shall provide qualified facilitators, all curriculum "
               "materials, and public liability insurance covering registered "
               "participants for the duration of the programme.")

    y = _numbered_section(c, MARGIN, y, '4.', 'Confidentiality', BRIGHT_BRAND)
    y = _prose(c, MARGIN, y,
               "Each party shall treat the personal data of participants "
               "disclosed under this MoU in accordance with the Personal Data "
               "Protection Act 2012 and shall not disclose such data to any "
               "third party save as required for the delivery of the Programme.")

    y = _numbered_section(c, MARGIN, y, '5.', 'Termination', BRIGHT_BRAND)
    y = _prose(c, MARGIN, y,
               "Either party may terminate this MoU by giving thirty (30) days' "
               "prior written notice to the other party.")

    y = _numbered_section(c, MARGIN, y, '6.', 'Executed by', BRIGHT_BRAND)
    y -= 4 * mm
    y = _signature_grid(c, MARGIN, y, PAGE_W - 2 * MARGIN, [
        {'printed_name': 'Nurul Aisyah',
         'role': 'Programmes Lead, Brightpath Learning LLP',
         'id_label': 'UEN', 'id_value': 'T18LL0042K',
         'date': '2026-05-15'},
        {'printed_name': 'Cheryl Ong',
         'role': 'Programme Director, PartnerConnect Events',
         'id_label': 'UEN', 'id_value': '201811234K',
         'date': '2026-05-15'},
    ])

    # small provider stamp overlay near the left signature
    _stamp(c, MARGIN + 55 * mm, y + 22 * mm,
           'BRIGHTPATH LLP', 'PROVIDER · 2026',
           brand=BRIGHT_BRAND, r=13 * mm)

    _footer_legal(c, [
        'Brightpath Learning LLP  ·  UEN T18LL0042K  ·  15 Beach Road #04-08, Singapore 189677  ·  brightpath.edu.sg',
        'This MoU has been reviewed by counsel (M/s Rajah & Tan, Ref. RT-24-1128) and is executed in two originals; one to be retained by each party.',
    ])
    _watermark(c); c.save()

    # ---- Brightpath contact sheet
    write_xlsx(d / 'brightpath_contacts.xlsx', 'Contact Sheet',
               title='Brightpath Learning · Primary Contact',
               subtitle='Extracted from internal partner directory · last verified 2026-05-30',
               title_bg='2E7D32',
               rows=[
                   ('Doc ID', 'doc_004'),
                   ('Organisation', 'Brightpath Learning LLP'),
                   ('Contact Name', 'Nurul Aisyah'),
                   ('Email', 'aisyah@brightpath.edu.sg'),
                   ('Phone', '+65 9876 5432'),
               ],
               extra_rows=[
                   ('Role', 'Programmes Lead'),
                   ('Preferred Hours', 'Weekdays 9am - 6pm'),
                   ('Escalation Contact', 'ops@brightpath.edu.sg'),
               ])

    # ---- Cobalt registration
    write_xlsx(d / 'partner_registration_cobalt.xlsx', 'Partner Registration',
               title='PartnerConnect · Vendor Registration',
               subtitle='Registration Portal partners.event.sg  ·  Series 2026-B',
               title_bg='0B3D91',
               rows=[
                   ('Doc ID', 'doc_005'),
                   ('Organisation', 'Cobalt Studio'),
                   ('Contact Name', 'Tan Mei Ling'),
                   ('Email', 'meiling@cobaltstudio.co'),
                   ('Phone', '+65 6555 2021'),
                   ('UEN', '53210987B'),
                   ('RSVP Status', 'Tentative'),
               ],
               extra_rows=[
                   ('Submitted', '2026-05-14 15:20 SGT'),
                   ('Reviewed by', 'partners@event.sg'),
                   ('Portal Ref', 'PC-VR-2026-04-2318'),
               ])

    # ---- Cobalt email
    write_txt(d / 'cobalt_email.txt', [
        'From: meiling@cobaltstudio.co',
        'To: partners@event.sg',
        'Sent: 2026-05-30 11:24 SGT',
        'Subject: Re: Partner event — confirming contact details',
        '',
        ('Doc ID', 'doc_006'),
        '',
        'Hi team,',
        '',
        "Thanks for the follow-up. Confirming my details for the event as",
        'below. The phone number on the registration form was our office',
        "line, but the mobile below is the fastest way to reach me on the",
        'day of the event.',
        '',
        '--',
        'Cobalt Studio',
        'Tan Mei Ling  ·  Creative Director',
        'meiling@cobaltstudio.co',
        '+65 6555 2020',
        '25A Cantonment Road #03-01, Singapore 089745',
        'UEN 53210987B',
        '--',
        '',
        'Looking forward to it. Ping me if anything else is needed on our',
        'end before the walk-through.',
        '',
        'Mei Ling',
    ])

    write_json(d / 'ground_truth.json', {
        'planted_conflicts': [
            {'field': 'Email', 'entity': 'Lim Wei Jie',
             'per_doc': {'doc_001': 'weijie.lim@acmerobotics.sg', 'doc_002': 'wjlim@acme.com.sg'},
             'correct': 'weijie.lim@acmerobotics.sg'},
            {'field': 'Phone', 'entity': 'Cobalt Studio',
             'per_doc': {'doc_005': '+65 6555 2021', 'doc_006': '+65 6555 2020'},
             'correct': '+65 6555 2020', 'note': 'doc_005 transcription typo'},
        ],
        'planted_gaps': [
            {'field': 'RSVP status', 'entity': 'Brightpath Learning LLP',
             'reason': 'not captured in either doc'},
        ],
        'alias_variations': [
            {'canonical': 'Acme Robotics Pte Ltd', 'aliases': ['ACME Robotics']},
        ],
        'sensitive_spans': [
            {'value': 'Lim Wei Jie', 'type': 'PERSON'},
            {'value': 'Nurul Aisyah', 'type': 'PERSON'},
            {'value': 'Tan Mei Ling', 'type': 'PERSON'},
            {'value': 'weijie.lim@acmerobotics.sg', 'type': 'EMAIL'},
            {'value': 'wjlim@acme.com.sg', 'type': 'EMAIL'},
            {'value': 'aisyah@brightpath.edu.sg', 'type': 'EMAIL'},
            {'value': 'meiling@cobaltstudio.co', 'type': 'EMAIL'},
            {'value': '+65 6123 4567', 'type': 'PHONE'},
            {'value': '+65 9876 5432', 'type': 'PHONE'},
            {'value': '+65 6555 2020', 'type': 'PHONE'},
            {'value': '+65 6555 2021', 'type': 'PHONE'},
            {'value': '201912345A', 'type': 'UEN'},
            {'value': 'T18LL0042K', 'type': 'UEN'},
            {'value': '53210987B', 'type': 'UEN'},
        ],
    })


# ============================================================
# PATIENT PILE
# ============================================================

def build_patient():
    d = OUT / 'patient_intake'
    d.mkdir(parents=True, exist_ok=True)

    # ---- Intake form (Rajesh) — clinical band header + zebra-striped
    #      particulars (distinct from KYC's lettered section bands)
    c = canvas.Canvas(str(d / 'intake_kumar.pdf'), pagesize=A4)
    y = _band_header(c, HEART_BRAND, 'National Heart Centre',
                     'Outpatient Cardiology · MoH-accredited healthcare institution',
                     '5 Hospital Drive, Singapore 169609',
                     'Tel +65 6704 8000  ·  nhcs.com.sg  ·  UEN 199801148C',
                     meta=[('MRN', 'NHC-8829471'),
                           ('Received', '2026-06-15'),
                           ('Consultation', '2026-07-02 10:30')])

    y = _title(c, MARGIN, y, 'Outpatient Intake Form',
               'Cardiology consultation · Ministry of Health National Standards for Medical Records')

    # thin clinical prose paragraph
    y = _prose(c, MARGIN, y,
               "This form captures the particulars submitted by the patient at "
               "registration. All fields marked as required under the Ministry of "
               "Health National Standards for Medical Records must be completed "
               "before the consultation begins.")

    # small caps section header (no shaded band, no letter box)
    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'PATIENT PARTICULARS')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 5 * mm
    y = _zebra_fields(c, MARGIN, y, [
        ('Doc ID', 'doc_001'),
        ('MRN', 'NHC-8829471'),
        ('Patient Name', 'Rajesh Kumar'),
        ('Date of Birth', '1986-03-14'),
        ('NRIC / FIN', 'S8612345Z'),
        ('Sex', 'Male'),
        ('Nationality', 'Singaporean'),
        ('Race', 'Indian'),
        ('Contact', '+65 9821 3345'),
        ('Language', 'English, Tamil'),
    ], label_w=48 * mm)

    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'CLINICAL DETAILS')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 5 * mm
    y = _zebra_fields(c, MARGIN, y, [
        ('Chief Complaint', 'Occasional chest discomfort; on treatment for HTN'),
        ('Referring Clinician', 'Dr Lee Wai Meng (MCR 32194B)'),
        ('Referring Clinic', 'Clinic ABC, 218 East Coast Road #02-33'),
        ('Reason for Referral', 'Further cardiology assessment'),
        ('Allergies', 'Penicillin'),
    ], label_w=48 * mm)

    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'INSURANCE & CONSENT')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 5 * mm
    y = _zebra_fields(c, MARGIN, y, [
        ('Insurance', 'Prudential PRUShield'),
        ('Policy Number', 'PRU-009281'),
        ('Effective From', '2022-01-01'),
        ('Plan', 'Premier'),
        ('Consent Signed', ''),
        ('Consent Date', ''),
    ], label_w=48 * mm)
    y -= 2 * mm

    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'SIGNATURES')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 6 * mm
    y = _signature_grid(c, MARGIN, y, PAGE_W - 2 * MARGIN, [
        {'printed_name': 'Rajesh Kumar',
         'role': 'Patient',
         'id_label': 'NRIC', 'id_value': 'S8612345Z',
         'date': '(pending signature)'},
        {'printed_name': 'Farah Nisha bte Iskandar',
         'role': 'Reception, Cardiology OP',
         'id_label': 'Staff ID', 'id_value': 'NHC-RC-4412',
         'date': '2026-06-15'},
    ])

    _footer_legal(c, [
        'National Heart Centre Singapore Pte Ltd  ·  Company Reg No. 199801148C  ·  5 Hospital Drive, Singapore 169609',
        'This form is a medical record. Unauthorised disclosure is an offence under the Personal Data Protection Act 2012 and the Private Hospitals and Medical Clinics Act.',
    ])
    _watermark(c); _page_num(c, 1, 1); c.save()

    # ---- Referral email (txt)
    write_txt(d / 'referral_kumar.txt', [
        'From: drlee@clinicabc.sg',
        'To: cardiology@bighospital.sg',
        'Cc: patient_records@bighospital.sg',
        'Sent: 2026-06-15 10:32 SGT',
        'Subject: Referral — R. Kumar · further cardiology assessment',
        'Attachments: ecg_2026-06-10.pdf, lab_2026-06-12.pdf',
        '',
        ('Doc ID', 'doc_002'),
        '',
        'Dear colleague,',
        '',
        'Please see the following patient for further cardiology assessment.',
        'He has been on my books for eighteen months for hypertension,',
        'currently on Losartan 50mg daily. At his recent review, his ECG',
        'showed occasional T-wave inversions in the lateral leads that I',
        'would like a specialist opinion on.',
        '',
        ('Patient', 'R. Kumar'),
        ('Date of Birth', '1986-04-14'),
        ('NRIC / FIN', 'S8612345Z'),
        '',
        ('Referring Clinician', 'Dr Lee Wai Meng'),
        ('MCR', '32194B'),
        ('Reason', 'Further cardiology assessment'),
        '',
        'I am attaching his most recent lab requisition and ECG report',
        'separately for your reference. Kindly copy the reception team on',
        'the appointment confirmation so I can update the patient.',
        '',
        'Regards,',
        'Dr Lee Wai Meng, MBBS (Sing), FRACGP',
        'Clinic ABC, 218 East Coast Road #02-33, Singapore 428917',
        'Tel +65 6440 2211  ·  MCR 32194B',
    ])

    # ---- Lab requisition (Innoquest) — two-column top + ruled tests table
    c = canvas.Canvas(str(d / 'lab_kumar.pdf'), pagesize=A4)
    y = _letterhead(c, INNO_BRAND, 'Innoquest Diagnostics',
                    'Clinical laboratory services · MOH-licensed',
                    '63 Hillview Avenue #06-16, Singapore 669569',
                    'Tel +65 6580 8600  ·  innoquest.sg  ·  UEN 199304815K',
                    meta=[
                        ('Lab Serial', 'LABREQ-44120'),
                        ('Received', '2026-06-16 · 08:12'),
                        ('Priority', 'Routine'),
                        ('Container', 'SST · Purple · Grey'),
                    ])

    y = _title(c, MARGIN, y, 'Lab Requisition',
               'Requesting Clinician: Dr Lee Wai Meng · Clinic ABC · MCR 32194B')

    # patient + clinician side-by-side
    y = _two_boxes(c, y,
                   left_title='Patient particulars',
                   left_rows=[
                       ('Doc ID', 'doc_003'),
                       ('Patient', 'Rajesh Kumar'),
                       ('NRIC / FIN', 'S8612345Z'),
                       ('Date of Birth', '1986-03-14'),
                       ('Sex', 'Male'),
                       ('Age', '40 y'),
                   ],
                   right_title='Referring clinician',
                   right_rows=[
                       ('Clinician', 'Dr Lee Wai Meng'),
                       ('MCR', '32194B'),
                       ('Clinic', 'Clinic ABC'),
                       ('Address', '218 East Coast Rd #02-33'),
                       ('Tel', '+65 6440 2211'),
                       ('Fax', '+65 6440 2212'),
                   ],
                   brand=INNO_BRAND)

    y = _prose(c, MARGIN, y,
               "Clinical details: 40M with treated hypertension. ECG shows occasional "
               "lateral-lead T-wave inversions. Requesting a lipid panel, CRP and HbA1c "
               "as part of cardiovascular risk stratification prior to specialist "
               "review. Fasting required for the lipid panel.")
    y -= 1 * mm

    c.setFillColor(INK); c.setFont('Helvetica-Bold', 9.5)
    c.drawString(MARGIN, y, 'Tests requested')
    y -= 4 * mm
    y = _ruled_table(c, MARGIN, y,
                     headers=['Panel', 'Test code', 'Container',
                              'Fasting', 'Turnaround'],
                     rows=[
                         ['Lipid panel (fasting)', 'LP-04', 'SST 5 mL',
                          'Yes', '24 h'],
                         ['C-reactive protein', 'CRP-01', 'SST 3 mL',
                          'No', '24 h'],
                         ['HbA1c', 'HB1-05', 'EDTA 3 mL',
                          'No', '48 h'],
                     ],
                     col_widths=[54 * mm, 22 * mm, 30 * mm,
                                 20 * mm, 28 * mm])
    y -= 2 * mm

    # signature + stamp
    y = _signature_grid(c, MARGIN, y, PAGE_W - 2 * MARGIN, [
        {'printed_name': 'Dr Lee Wai Meng',
         'role': 'Requesting clinician',
         'id_label': 'MCR', 'id_value': '32194B',
         'date': '2026-06-15'},
        {'printed_name': 'Innoquest specimen desk',
         'role': 'Received and accessioned',
         'id_label': 'Lab ID', 'id_value': 'IQ-DESK-07',
         'date': '2026-06-16 08:12'},
    ])
    _stamp(c, PAGE_W - MARGIN - 20 * mm, y + 24 * mm,
           'INNOQUEST', 'SPECIMEN · ACCEPTED',
           brand=INNO_BRAND, r=13 * mm)

    _footer_legal(c, [
        'Innoquest Diagnostics Pte Ltd  ·  UEN 199304815K  ·  63 Hillview Avenue #06-16, Singapore 669569  ·  innoquest.sg',
        'Please phone the specimen desk on +65 6580 8600 for urgent requisitions or STAT collection. All specimens are handled under ISO 15189 accreditation.',
    ])
    _watermark(c); c.save()

    # ---- Intake (Chloe)
    c = canvas.Canvas(str(d / 'intake_ng.pdf'), pagesize=A4)
    y = _band_header(c, HEART_BRAND, 'National Heart Centre',
                     'Outpatient Cardiology · MoH-accredited healthcare institution',
                     '5 Hospital Drive, Singapore 169609',
                     'Tel +65 6704 8000  ·  nhcs.com.sg  ·  UEN 199801148C',
                     meta=[('MRN', 'NHC-8829504'),
                           ('Received', '2026-06-10'),
                           ('Consultation', '2026-06-28 09:15')])

    y = _title(c, MARGIN, y, 'Outpatient Intake Form',
               'Cardiology consultation · Ministry of Health National Standards for Medical Records')

    y = _prose(c, MARGIN, y,
               "This form captures the particulars submitted by the patient at "
               "registration. All fields marked as required under the Ministry of "
               "Health National Standards for Medical Records must be completed "
               "before the consultation begins.")

    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'PATIENT PARTICULARS')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 5 * mm
    y = _zebra_fields(c, MARGIN, y, [
        ('Doc ID', 'doc_004'),
        ('MRN', 'NHC-8829504'),
        ('Patient Name', 'Chloe Ng'),
        ('Date of Birth', '1994-11-02'),
        ('NRIC / FIN', 'S9456781D'),
        ('Sex', 'Female'),
        ('Nationality', 'Singaporean'),
        ('Race', 'Chinese'),
        ('Contact', '+65 9662 8114'),
        ('Language', 'English, Mandarin'),
    ], label_w=48 * mm)

    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'CLINICAL DETAILS')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 5 * mm
    y = _zebra_fields(c, MARGIN, y, [
        ('Chief Complaint', 'Palpitations on exertion x 6 weeks'),
        ('Referring Clinician', 'Dr Kavitha S (MCR 40821E)'),
        ('Referring Clinic', 'Novena Cardiology Clinic, 10 Sinaran Drive'),
        ('Reason for Referral', 'Palpitation workup'),
        ('Allergies', ''),
    ], label_w=48 * mm)

    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'INSURANCE & CONSENT')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 5 * mm
    y = _zebra_fields(c, MARGIN, y, [
        ('Insurance', 'AIA HealthShield Gold Max'),
        ('Policy Number', 'AIA-553102'),
        ('Effective From', '2024-01-01'),
        ('Plan', 'Standard'),
        ('Consent Signed', 'Yes (2026-06-10)'),
        ('Consent Date', '2026-06-10'),
    ], label_w=48 * mm)

    c.setFillColor(HEART_BRAND); c.setFont('Helvetica-Bold', 9)
    c.drawString(MARGIN, y, 'SIGNATURES')
    _rule(c, MARGIN, y - 1.5 * mm, PAGE_W - MARGIN, colour=HEART_BRAND, w=0.6)
    y -= 6 * mm
    y = _signature_grid(c, MARGIN, y, PAGE_W - 2 * MARGIN, [
        {'printed_name': 'Chloe Ng',
         'role': 'Patient',
         'id_label': 'NRIC', 'id_value': 'S9456781D',
         'date': '2026-06-10'},
        {'printed_name': 'Farah Nisha bte Iskandar',
         'role': 'Reception, Cardiology OP',
         'id_label': 'Staff ID', 'id_value': 'NHC-RC-4412',
         'date': '2026-06-10'},
    ])

    _footer_legal(c, [
        'National Heart Centre Singapore Pte Ltd  ·  Company Reg No. 199801148C  ·  5 Hospital Drive, Singapore 169609',
        'This form is a medical record. Unauthorised disclosure is an offence under the Personal Data Protection Act 2012.',
    ])
    _watermark(c); _page_num(c, 1, 1); c.save()

    # ---- Insurance card (AIA) — landscape card mockup on A4
    c = canvas.Canvas(str(d / 'insurance_ng.pdf'), pagesize=A4)
    y = _letterhead(c, AIA_BRAND, 'AIA Singapore',
                    'AIA HealthShield Gold Max · Integrated Shield Plan',
                    '1 Robinson Road, AIA Tower, Singapore 048542',
                    'Tel 1800 248 8000  ·  aia.com.sg  ·  UEN 201106386R',
                    meta=[
                        ('Policy No.', 'AIA-553120'),
                        ('Card No.', 'AIA-CARD-4451-9932'),
                        ('Group', 'IND-STD-A'),
                        ('Effective', '2024-01-01'),
                    ])

    y = _title(c, MARGIN, y, 'Health Insurance Membership Card',
               'Card image scanned by the clinic reception on registration · Photocopy for verification')

    y = _prose(c, MARGIN, y,
               "Please present this card at the point of service. Cashless "
               "hospitalisation is available at all AIA panel institutions. For "
               "pre-authorisation queries or claim status, quote the policy number "
               "printed below and the member's NRIC.")

    # landscape card mockup
    cw = 150 * mm; ch = 90 * mm
    cx = (PAGE_W - cw) / 2; cy = y - ch - 6 * mm
    # background card
    c.setStrokeColor(HAIR); c.setLineWidth(0.6)
    c.setFillColor(HexColor('#FFFFFF'))
    c.roundRect(cx, cy, cw, ch, 5 * mm, fill=1, stroke=1)
    # brand strip
    c.setFillColor(AIA_BRAND)
    c.rect(cx, cy + ch - 18 * mm, cw, 18 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont('Helvetica-Bold', 15)
    c.drawString(cx + 6 * mm, cy + ch - 10 * mm, 'AIA Singapore')
    c.setFont('Helvetica', 9)
    c.drawString(cx + 6 * mm, cy + ch - 14.5 * mm,
                 'HealthShield Gold Max A')
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(cx + cw - 6 * mm, cy + ch - 10 * mm,
                      'MEMBER CARD')
    c.setFont('Helvetica', 8)
    c.drawRightString(cx + cw - 6 * mm, cy + ch - 14.5 * mm,
                      'Integrated Shield Plan')
    # data grid
    c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 8)
    c.drawString(cx + 6 * mm, cy + ch - 26 * mm, 'CARDHOLDER')
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 14)
    c.drawString(cx + 6 * mm, cy + ch - 32 * mm, 'Chloe Ng')

    c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 8)
    c.drawString(cx + 6 * mm, cy + 44 * mm, 'POLICY NUMBER')
    c.setFillColor(INK); c.setFont('Helvetica-Bold', 12)
    c.drawString(cx + 6 * mm, cy + 38 * mm, 'AIA-553120')

    c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 8)
    c.drawString(cx + 78 * mm, cy + 44 * mm, 'EFFECTIVE FROM')
    c.setFillColor(INK); c.setFont('Helvetica', 11)
    c.drawString(cx + 78 * mm, cy + 38 * mm, '2024-01-01')

    c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 8)
    c.drawString(cx + 6 * mm, cy + 24 * mm, 'PLAN')
    c.setFillColor(INK); c.setFont('Helvetica', 11)
    c.drawString(cx + 6 * mm, cy + 18 * mm, 'Health Shield Gold Max A')

    c.setFillColor(MUTED); c.setFont('Helvetica-Bold', 8)
    c.drawString(cx + 78 * mm, cy + 24 * mm, 'GROUP')
    c.setFillColor(INK); c.setFont('Helvetica', 11)
    c.drawString(cx + 78 * mm, cy + 18 * mm, 'IND-STD-A')

    # small print + emergency
    c.setFillColor(FAINT); c.setFont('Helvetica-Oblique', 6.8)
    c.drawString(cx + 6 * mm, cy + 6 * mm,
                 '24-hour emergency: 1800 248 8000  ·  claims@aia.com.sg  ·  aia.com.sg/claim')

    # transcription box below the card
    y = cy - 6 * mm
    y = _boxed_panel(c, MARGIN, y, PAGE_W - 2 * MARGIN,
                     'Transcription',
                     [
                         ('Doc ID', 'doc_005'),
                         ('Cardholder', 'Chloe Ng'),
                         ('Policy Number', 'AIA-553120'),
                         ('Plan', 'Health Shield Gold Max A'),
                         ('Effective From', '2024-01-01'),
                         ('Group', 'IND-STD-A'),
                     ],
                     AIA_BRAND, label_w=44 * mm)

    _footer_legal(c, [
        'AIA Singapore Private Limited  ·  UEN 201106386R  ·  1 Robinson Road, AIA Tower, Singapore 048542  ·  aia.com.sg',
        'This card remains the property of AIA and must be returned upon request. Loss of the physical card should be reported to member services immediately.',
    ])
    _watermark(c); c.save()

    write_json(d / 'ground_truth.json', {
        'planted_conflicts': [
            {'field': 'Date of birth', 'entity': 'Rajesh Kumar',
             'per_doc': {'doc_001': '1986-03-14', 'doc_002': '1986-04-14'},
             'correct': '1986-03-14', 'note': 'doc_002 month transposition'},
            {'field': 'Policy number', 'entity': 'Chloe Ng',
             'per_doc': {'doc_004': 'AIA-553102', 'doc_005': 'AIA-553120'},
             'correct': 'AIA-553102', 'note': 'doc_005 final digits transposed'},
        ],
        'planted_gaps': [
            {'field': 'Consent signed', 'entity': 'Rajesh Kumar',
             'reason': 'intake unsigned'},
            {'field': 'Allergies', 'entity': 'Chloe Ng',
             'reason': 'field left blank on intake'},
        ],
        'alias_variations': [
            {'canonical': 'Rajesh Kumar', 'aliases': ['R. Kumar']},
        ],
        'sensitive_spans': [
            {'value': 'Rajesh Kumar', 'type': 'PERSON'},
            {'value': 'R. Kumar', 'type': 'PERSON'},
            {'value': 'Chloe Ng', 'type': 'PERSON'},
            {'value': 'S8612345Z', 'type': 'NRIC'},
            {'value': 'S9456781D', 'type': 'NRIC'},
            {'value': 'PRU-009281', 'type': 'POLICY'},
            {'value': 'AIA-553102', 'type': 'POLICY'},
            {'value': 'AIA-553120', 'type': 'POLICY'},
            {'value': 'LABREQ-44120', 'type': 'SERIAL'},
            {'value': '1986-03-14', 'type': 'DATE'},
            {'value': '1986-04-14', 'type': 'DATE'},
            {'value': '1994-11-02', 'type': 'DATE'},
        ],
    })


# ============================================================
# BITMAP SAMPLES (Real-pile upload demo)
# ============================================================

def build_samples():
    d = OUT / 'samples'
    d.mkdir(parents=True, exist_ok=True)

    _render_id_card_png(d / 'nric_scan_chloe.png',
                        name='Chloe Ng',
                        nric='S9456781D',
                        dob='02 Nov 1994')

    _render_scan_jpeg(d / 'utility_bill_priya.jpg',
                      title='SP Group · Utility Bill',
                      brand=(0, 169, 206),
                      rows=[
                          ('Doc ID', 'doc_001'),
                          ('Account Holder', 'Priya Suresh'),
                          ('NRIC', 'S8712345B'),
                          ('Service Address', 'Blk 44 Toa Payoh Lor 5 #08-217'),
                          ('', 'Singapore 310044'),
                          ('Billing Period', '2026-06-01 to 2026-06-30'),
                          ('Amount Due', 'SGD 92.10'),
                      ])

    _render_signed_letter_jpeg(
        d / 'employment_letter_priya.jpg',
        header='TechFlow Consulting',
        tagline='Global consulting services  ·  #22-01 Suntec Tower 5, Singapore 038985',
        brand=(11, 61, 145),
        body_paragraphs=[
            'Date: 2026-06-04',
            '',
            'To whom it may concern,',
            '',
            'This letter certifies that Priya Suresh (NRIC S8712345B) has '
            'been employed with TechFlow Consulting Pte Ltd since '
            '2019-03-11 as a Senior Consultant. Her current monthly gross '
            'salary is SGD 9,800.',
            'She may contact HR at hr@techflow.sg or +65 6812 4400 for any '
            'verification queries.',
        ],
        sign_off_name='Emily Tan',
        sign_off_title='Head of People, TechFlow Consulting Pte Ltd'
    )

    (d / 'README.md').write_text(
        '# scanned-doc samples\n\n'
        'A small set of image documents (PNG/JPEG) used to demo the Real-pile '
        'upload flow. These are NOT part of any pile\'s ground truth; drag '
        'them into the "Your pile" tab in the UI to see Gemma extract the '
        'fields live and Presidio redact identifiers.\n\n'
        '- `nric_scan_chloe.png` — Singapore NRIC-shaped ID card photo\n'
        '- `utility_bill_priya.jpg` — scanned utility bill with a rubber-stamp receipt marker\n'
        '- `employment_letter_priya.jpg` — signed employment verification letter\n'
    )


def build_readme():
    (OUT / 'README.md').write_text(
        '# paperflow synthetic data\n\n'
        'Three demo piles for the paperflow reconciler. Regenerate with:\n\n'
        '    python3 generate_synthetic.py\n\n'
        '## piles\n\n'
        '- `kyc_onboarding/` — 4 documents (KYC form, utility bill, NRIC scan, email declaration)\n'
        '- `partner_collation/` — 6 documents (registration sheets, MoU, business card, email)\n'
        '- `patient_intake/` — 5 documents (intake forms, referral letter, lab requisition, insurance card)\n\n'
        '`samples/` also contains scanned-image documents (PNG/JPEG) for the Real-pile upload demo.\n\n'
        '## visual style\n\n'
        'Every PDF is composed from the same primitives (letterhead with a meta '
        'bar of document references, section bands, framed data panels, ruled '
        'tables, signature grids, barcodes and rubber stamps) so the pages read '
        'like real business/clinical documents rather than a stripped-down '
        'mock-up. Every planted VALUE stays identical across regenerations to '
        "keep the scorer's ground-truth stable.\n\n"
        '## eval\n\n'
        'Each pile ships a `ground_truth.json` with:\n\n'
        '- `planted_conflicts` — same field, different value per doc; scorer checks the reconciler picks `correct`.\n'
        '- `planted_gaps` — required fields missing everywhere in the pile.\n'
        '- `alias_variations` — same entity, different surface form; scorer checks merge not flag.\n'
        '- `sensitive_spans` — every value redaction recall must catch.\n\n'
        'All names, IDs, addresses, phones, emails and policies are fictional.\n'
    )


def main():
    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True)
    build_kyc()
    build_partner()
    build_patient()
    build_samples()
    build_readme()
    print(f'Generated synthetic data at {OUT}/')
    for sub in sorted(OUT.iterdir()):
        if sub.is_dir():
            files = sorted(sub.iterdir())
            print(f'  {sub.name}/  ({len(files)} files)')
            for f in files:
                print(f'    {f.name}')


if __name__ == '__main__':
    main()
